"""Excel 填报模板样式与数据验证（导出与模板生成共用）。

与 docs/项目填报模板.xlsx 的格式保持一致：
- R1 标题（合并居中）
- R2 表头（深色底白字）
- R3 起数据行（边框、字体、关键列居中）
- 冻结 A3
- 数据验证：类目 / 市场 / 阶段类型 / 阶段状态 下拉 + 进度 0-100 + 日期范围
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- 表头与列宽（与 docs/项目填报模板.xlsx 一致，14 列） ----------
HEADERS = [
    "项目编号",  # A
    "项目类目",  # B
    "项目名称",  # C
    "项目负责人",  # D
    "市场",      # E
    "阶段类型",  # F
    "计划开始",  # G
    "计划结束",  # H
    "实际开始",  # I
    "实际结束",  # J
    "阶段负责人",  # K
    "阶段状态",  # L
    "阶段进度",  # M
    "备注",      # N
]

COL_WIDTHS = [12, 10, 30, 12, 10, 16, 13, 13, 13, 13, 18, 10, 8, 24]

# ---------- 样式 ----------
HEADER_FILL = PatternFill(start_color="001529", end_color="001529", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="001529")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")

# ---------- 下拉选项（与 docs/项目填报模板.xlsx 一致） ----------
CATEGORIES = '"新需求,量产,定制,改造"'
MARKETS = '"拉美区,西欧区,东欧区,中东区,亚太区,土耳其区,非洲区,北美区,OEM业务部"'
PHASE_TYPES = '"需求评估,配置评估,模块选型,工业设计,结构设计,样机打样,联调测试,交付"'
PHASE_STATUSES = '"未开始,进行中,已完成,延期,已搁置"'


def style_sheet(ws, data_rows: int = 200) -> None:
    """把工作表套用为模板样式（标题/表头/数据行/冻结/数据验证）。

    - 假定表头在第 2 行、数据从第 3 行开始（调用方负责写标题与表头内容）
    - data_rows：数据行数（决定验证范围，默认 200 与模板一致）
    """
    COLS = len(HEADERS)
    LAST_COL = get_column_letter(COLS)

    # 列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结表头
    ws.freeze_panes = "A3"

    # 第 1 行：标题
    ws.merge_cells(f"A1:{LAST_COL}1")
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 第 2 行：表头
    ws.row_dimensions[2].height = 22
    for col in range(1, COLS + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 数据行
    for r in range(3, data_rows + 3):
        ws.row_dimensions[r].height = 22
        for c in range(1, COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if c in (1, 2, 4, 5, 6, 7, 8, 9, 10, 12, 13):
                cell.alignment = CENTER

    last = f"{data_rows + 2}"

    # 类目 (B)
    dv = DataValidation(type="list", formula1=CATEGORIES, allow_blank=True)
    dv.error = "请从下拉列表中选择"
    ws.add_data_validation(dv); dv.add(f"B3:B{last}")

    # 市场 (E)
    dv = DataValidation(type="list", formula1=MARKETS, allow_blank=True)
    dv.error = "请从下拉列表中选择市场区域"
    ws.add_data_validation(dv); dv.add(f"E3:E{last}")

    # 阶段类型 (F)
    dv = DataValidation(type="list", formula1=PHASE_TYPES, allow_blank=True)
    dv.error = "请从下拉列表中选择阶段类型"
    ws.add_data_validation(dv); dv.add(f"F3:F{last}")

    # 阶段状态 (L)
    dv = DataValidation(type="list", formula1=PHASE_STATUSES, allow_blank=True)
    dv.error = "请从下拉列表中选择阶段状态"
    ws.add_data_validation(dv); dv.add(f"L3:L{last}")

    # 阶段进度 (M) — 0-100
    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="100")
    dv.error = "进度请填 0-100"
    ws.add_data_validation(dv); dv.add(f"M3:M{last}")

    # 日期 (G-J)
    dv = DataValidation(type="date", operator="greaterThan", formula1="2024-01-01")
    dv.error = "请填入有效日期"
    ws.add_data_validation(dv); dv.add(f"G3:J{last}")
