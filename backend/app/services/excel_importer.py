"""Excel 导入器：解析项目进度 Excel，写入数据库。

支持两种格式（按表头自动识别）：
1. 新格式（14 列，docs/项目填报模板.xlsx，推荐）：
   项目编号 | 项目类目 | 项目名称 | 项目负责人 | 市场 | 阶段类型 |
   计划开始 | 计划结束 | 实际开始 | 实际结束 | 阶段负责人 | 阶段状态 | 阶段进度 | 备注
2. 旧格式（8 列，历史源文件）：
   项目编号 | 项目类目 | 项目名称 | 负责人 | 计划开始 | 计划结束 | 状态 | 交接人

通用规则（解析层以 docs/EXCEL_PARSE_COMPAT.md 为权威）：
- 数据从第 3 行开始（前 2 行表头）
- 单元格清洗 clean_cell（§2.3）：NFKC 全角→半角、去不可见字符、压缩连续空白，应用于所有文本列
- 行判定（§2.1，阶段类型优先）：阶段类型列可解析 → 阶段行；编号纯数字+名称非空 → 项目行；
  编号 1-1 / 1.1 → 阶段行；其他有内容 → 跳过 + 警告（§2.7 可见化）
- 日期多格式链（§2.4）：Excel 序列号（1..2958465）→ YYYY[-/.]M[-/.]D → YYYY年M月D日
  → 8/7 位纯数字 → 短格式 M-D（按当年解析 + 警告）；失败 → warning 设 NULL
- 进度统一 0-100（§2.5）：50% / 50％ / 0.5 / 50 → 50
- 多人字段：多空格/逗号分隔，拆分后逐人建/匹配 Resource

导入策略：全量重置（先删所有 Project/Phase/Dependency，清空 Resource，保留 Template）。
"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

import openpyxl
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.models import Dependency, Phase, Project, Resource, User, phase_assignee
from app.schemas.import_report import (
    ImportError,
    ImportExistingCounts,
    ImportIncomingCounts,
    ImportPreview,
    ImportReport,
    ImportWarning,
    PendingLinkPhase,
    PreviewProject,
    ProjectMatchInfo,
)

# ---------- 新格式（14 列，项目填报模板）列索引 ----------
_NEW_COL_CODE = 1
_NEW_COL_CATEGORY = 2
_NEW_COL_NAME = 3
_NEW_COL_OWNER = 4
_NEW_COL_MARKET = 5
_NEW_COL_PHASE_TYPE = 6
_NEW_COL_PLAN_START = 7
_NEW_COL_PLAN_END = 8
_NEW_COL_ACTUAL_START = 9
_NEW_COL_ACTUAL_END = 10
_NEW_COL_ASSIGNEES = 11
_NEW_COL_STATUS = 12
_NEW_COL_PROGRESS = 13
_NEW_COL_REMARK = 14

# ---------- 旧格式（8 列，历史源文件）列索引 ----------
_OLD_COL_CODE = 1
_OLD_COL_CATEGORY = 2
_OLD_COL_NAME = 3
_OLD_COL_OWNER = 4
_OLD_COL_PLAN_START = 5
_OLD_COL_PLAN_END = 6
_OLD_COL_STATUS = 7
_OLD_COL_HANDOVER = 8  # 旧格式读取但不再写入数据库（字段已废弃）

# 数据起始行（前 2 行表头）
_DATA_START_ROW = 3

# 项目编号正则：纯数字=项目；数字-数字 / 数字.数字=阶段（兼容新旧两种编号格式，§2.1 ③）
_PROJECT_CODE_RE = re.compile(r"^\d+$")
_PHASE_CODE_RE = re.compile(r"^\d+-\d+$")
_PHASE_DOT_CODE_RE = re.compile(r"^\d+\.\d+$")

# 阶段行父编号（'1-1' / '1.1' 的前段数字），用于 §2.2 归属校验
_PHASE_PARENT_RE = re.compile(r"^(\d+)[-.]\d+$")

# §2.3 清洗：NFKC 后仍残留的不可见字符（零宽/软换行/BOM/方向标记）
_INVISIBLE_RE = re.compile(r"[\u200b-\u200f\ufeff]")
# §2.3 清洗：连续空白压缩为单个空格
_MULTI_SPACE_RE = re.compile(r"\s+")

# §2.4 日期链正则（完整格式的月/日分隔符须一致；短格式 M[-/.]D 单分隔符）
_DATE_FULL_RE = re.compile(r"^(\d{4})([-/.])(\d{1,2})\2(\d{1,2})$")
_DATE_CN_RE = re.compile(r"^(\d{4})年(\d{1,2})月(\d{1,2})日?$")
_DATE_SHORT_RE = re.compile(r"^(\d{1,2})([-/.])(\d{1,2})$")

# Excel 日期序列号有效范围（1900-9999 年，评审处置 #5：越界不按序列号解析）
_SERIAL_MIN, _SERIAL_MAX = 1, 2_958_465

# 多人字段拆分正则：任意空白/中英文逗号/顿号
_PERSON_SPLIT_RE = re.compile(r"[\s,，、]+")


def _normalize_name(name: str) -> str:
    """名称比较键归一化（评审处置 #2 / #4）：NFKC + 去空白。

    用于合并导入的项目名比较键与阶段名映射表键，保证全角/半角、
    多空格等书写差异不影响命中。
    """
    return unicodedata.normalize("NFKC", name).strip()


# 阶段类型前缀（新格式 "P4 工业设计" → code=P4, name=工业设计）
_PHASE_TYPE_PREFIX_RE = re.compile(r"^(P[1-8])\s*(.*)$")

# 阶段名 → phase_type 映射表（旧格式/新格式阶段类型列缺省时的兜底）
# 基础部分来自 PROJECT_SPEC §5.3；扩充部分覆盖实际 Excel 里的阶段名变体。
# key 为阶段名（trim 后），value 为 phase_type。
# 查找走 _PHASE_NAME_LOOKUP（NFKC 归一化索引，评审处置 #4：全角括号键可匹配半角输入）。
PHASE_NAME_TO_TYPE: dict[str, str] = {
    # --- 基础映射（§5.3 原表）---
    "工业设计": "P4",
    "结构设计": "P5",
    "整机设计": "P5",
    "样机打样": "P6",
    "联调测试": "P7",
    "测试": "P7",
    "POC及投标": "P8",
    "需求分析": "P1",
    "需求评估": "P1",
    "配置评估": "P2",
    "模块选型": "P3",
    "直接投料": "P8",
    "图纸归档": "P5",
    "归档": "P5",
    "BOM制作与激活": "P8",
    "首批生产保障": "P8",
    "投料": "P8",
    "发货": "P8",
    # --- 扩充映射（实际 Excel 里的阶段名变体）---
    "测试与发货": "P8",          # = 交付
    "样机打样（1台）": "P6",     # = 样机打样
    "归档（归档后再投料）": "P8", # 量产阶段的归档后投料 = 交付（区别于设计阶段的"归档"P5）
    "直接投料，BOM制作与激活": "P8",  # = 交付
    "直接投料，激活时间": "P8",   # = 交付
    "交付": "P8",               # = 交付（直白词，§5.3 原表遗漏）
}

# NFKC 归一化索引（评审处置 #4）：clean_cell 后的输入（全角括号已被归一化为半角）
# 也能命中映射表内含全角括号的键，如「样机打样（1台）」
_PHASE_NAME_LOOKUP: dict[str, str] = {
    _normalize_name(k): v for k, v in PHASE_NAME_TO_TYPE.items()
}

# 新格式表头关键字（识别用）
_NEW_FORMAT_KEYS = ("项目编号", "阶段类型", "市场")
_OLD_FORMAT_KEYS = ("项目编号", "交接人")

# 进程内存：最近一次导入报告（供 GET /api/import/report 查询）
_last_report: ImportReport | None = None


def get_last_report() -> ImportReport | None:
    return _last_report


def _set_last_report(report: ImportReport | None) -> None:
    global _last_report
    _last_report = report


# ---------- 清洗工具函数 ----------

def clean_cell(value: Any) -> Any:
    """单元格清洗（§2.3，复制粘贴兼容）——仅处理文本单元格。

    1. NFKC 全角→半角（１→1、Ａ→A、（→(、．→.、％→%）
    2. 去不可见字符（零宽 \\u200b-\\u200f、BOM \\ufeff）
    3. 去首尾空白（含全角空格 \\u3000，NFKC 已把 \\u00a0 归为普通空格）
    4. 压缩连续空白为单个空格

    非字符串（None/datetime/数值）原样返回，交由各自解析器处理。
    """
    if not isinstance(value, str):
        return value
    text = unicodedata.normalize("NFKC", value)
    text = _INVISIBLE_RE.sub("", text)
    text = text.strip()
    text = _MULTI_SPACE_RE.sub(" ", text)
    return text


def _clean_str(value: Any) -> str:
    """clean_cell + str 化：任何单元格值 → 清洗后的字符串（None/空 → ''）。

    数值会先 str 化再清洗（进度 0 等有效值不会被丢弃）。
    """
    if value is None:
        return ""
    cleaned = clean_cell(value)
    return str(cleaned).strip() if cleaned is not None else ""


def _try_date(y: int, m: int, d: int) -> date | None:
    """构造 date，非法日期（如 2 月 30 日）返回 None 而非抛异常。"""
    try:
        return date(y, m, d)
    except ValueError:
        return None


def _warn_unparsed_date(text: str, row: int, sheet: str, field: str, report: ImportReport) -> None:
    """日期解析失败：记录 warning（保持既有行为：设空 + 可见）。"""
    report.warnings.append(ImportWarning(
        row=row, sheet=sheet, field=field,
        message=f"无法解析日期 '{text}'，已设为空",
    ))
    return None


def parse_cell_date(value: Any, row: int, sheet: str, field: str, report: ImportReport) -> date | None:
    """解析日期单元格（§2.4 多格式链，文本先经 clean_cell 清洗）。

    - datetime/date 对象 → 直接取 date
    - Excel 序列号 → 仅 1..2_958_465 有效（评审处置 #5）；数值型 7/8 位
      与文本链同序优先按 YYYYMMDD/YYYYMDD（20260629 / 2026629），非法才回退序列号或警告
    - 文本链（clean_cell 后依次尝试）：
      YYYY[-/.]M[-/.]D → YYYY年M月D日 → 8 位 YYYYMMDD / 7 位 YYYYMDD
      → 短格式 M[-/.]D（按当年解析 + 警告，决策 A）→ 数字文本序列号
    - 全部失败 → None + warning
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # 评审处置 #5 补全：数值型 7/8 位优先按 YYYYMMDD/YYYYMDD 解析（20260629 / 2026629），
        # 与文本链同序——否则区间内 7 位数字日期（2026629 < 2_958_465）会被当成
        # 序列号解析成 7447 年的荒谬日期；月/日非法（20261399）才回退序列号或警告
        digits = str(int(value))
        if len(digits) in (7, 8):
            y, rest = int(digits[:4]), digits[4:]
            d = _try_date(y, int(rest[:-2]), int(rest[-2:]))
            if d:
                return d
        # Excel 日期序列号（评审处置 #5：仅 1..2_958_465 按序列号解析，越界不猜）
        if _SERIAL_MIN <= value <= _SERIAL_MAX:
            try:
                return date(1899, 12, 30) + timedelta(days=int(value))
            except (ValueError, OverflowError):
                pass
        return _warn_unparsed_date(str(value), row, sheet, field, report)

    # 文本链（clean_cell 清洗）
    text = clean_cell(value)
    if not text:
        return None
    m = _DATE_FULL_RE.match(text)
    if m:
        d = _try_date(int(m.group(1)), int(m.group(3)), int(m.group(4)))
        if d:
            return d
    m = _DATE_CN_RE.match(text)
    if m:
        d = _try_date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        if d:
            return d
    if text.isdigit() and len(text) in (7, 8):
        # 8 位 YYYYMMDD / 7 位 YYYYMDD（20260629 / 2026629）
        y, rest = int(text[:4]), text[4:]
        d = _try_date(y, int(rest[:-2]), int(rest[-2:]))
        if d:
            return d
    m = _DATE_SHORT_RE.match(text)
    if m:
        # 短格式缺年份：按当年解析（决策 A）+ 警告
        year = date.today().year
        d = _try_date(year, int(m.group(1)), int(m.group(3)))
        if d:
            report.warnings.append(ImportWarning(
                row=row, sheet=sheet, field=field,
                message=f"日期'{text}'缺少年份，按{year}年解析",
            ))
            return d
    # 纯数字文本：Excel 序列号语义（范围内才解析）
    if text.isdigit() and _SERIAL_MIN <= int(text) <= _SERIAL_MAX:
        try:
            return date(1899, 12, 30) + timedelta(days=int(text))
        except (ValueError, OverflowError):
            pass

    return _warn_unparsed_date(text, row, sheet, field, report)


def parse_progress(value: Any) -> int | None:
    """进度解析统一 0-100（§2.5）：'50%' / '50％' / '50 %' / 0.5 / 50 → 50。

    - 百分号形式（NFKC 后半角 %）：取数值，夹取 0-100
    - 0-1 数值（含小数）：按比例 ×100（0.5 → 50；1 视为 100%）
    - 大于 1 的数值：按 0-100 原样，>100 夹取 100
    - 无法解析 → None（由调用方记警告并按状态推断）

    gantt API 层的 0-1 转换由现有 API 层负责，本函数只产 0-100。
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        if num <= 1:
            return int(round(num * 100))
        return max(0, min(100, int(round(num))))
    text = clean_cell(value)
    if not isinstance(text, str) or not text:
        return None
    if text == "/":
        return None
    pct = text.endswith("%")
    if pct:
        text = text[:-1].strip()
    try:
        num = float(text)
    except ValueError:
        return None
    if pct:
        return max(0, min(100, int(round(num))))
    if num <= 1:
        return int(round(num * 100))
    return max(0, min(100, int(round(num))))


def split_persons(value: Any) -> list[str]:
    """拆分多人字段，返回去重后的人员姓名列表（输入先经 clean_cell）。"""
    text_value = clean_cell(value) if isinstance(value, str) else value
    if text_value is None or text_value == "":
        return []
    # 清除换行，取有效内容
    text = str(text_value).replace("\n", " ").strip()
    parts = [p.strip() for p in _PERSON_SPLIT_RE.split(text) if p.strip()]
    # 去重保序
    seen: set[str] = set()
    result: list[str] = []
    for p in parts:
        if p not in seen and p != "/":  # "/" 在原表表示无
            seen.add(p)
            result.append(p)
    return result


def classify_row(code: Any, name: Any, phase_type_cell: Any = None) -> str:
    """行分类（§2.1，阶段类型优先）：'project' | 'phase' | 'skip'。

    ① 阶段类型列可解析（P1-P8 前缀或映射表命中）→ 'phase'（评审处置 #1，不依赖编号）
    ② 编号纯数字 且 名称非空 → 'project'
    ③ 编号 ^\\d+-\\d+$ 或 ^\\d+\\.\\d+$ → 'phase'（兼容新旧两种编号格式）
    ④ 编号/名称/阶段类型全空 → 'skip'（静默空行）
    其余有内容但无法识别 → 'skip'（由调用方记警告，§2.7）

    所有输入先经 clean_cell 清洗。
    """
    code_s = clean_cell(code)
    code_s = str(code_s).strip() if code_s is not None else ""
    name_s = clean_cell(name)
    name_s = str(name_s).strip() if name_s is not None else ""
    pt_cell = clean_cell(phase_type_cell) if phase_type_cell is not None else None

    # ④ 空行：三类关键单元格全空
    if not code_s and not name_s and not pt_cell:
        return "skip"

    # ① 阶段类型优先（评审处置 #1）
    if pt_cell and parse_phase_type_cell(pt_cell) is not None:
        return "phase"

    # ② 编号纯数字 + 名称非空 → 项目行
    if _PROJECT_CODE_RE.match(code_s) and name_s:
        return "project"

    # ③ 阶段编号两种格式
    if _PHASE_CODE_RE.match(code_s) or _PHASE_DOT_CODE_RE.match(code_s):
        return "phase"

    # 其余有内容但无法识别 → skip（调用方记警告）
    return "skip"


def map_phase_type(name: str, row: int, sheet: str, report: ImportReport) -> str | None:
    """阶段名 → phase_type（NFKC 归一化索引，评审处置 #4）。无法映射的记 error 并返回 None。"""
    phase_type = _PHASE_NAME_LOOKUP.get(_normalize_name(name.strip()))
    if phase_type is None:
        report.errors.append(ImportError(
            row=row, sheet=sheet, field="name",
            message=f"阶段名'{name}'无法映射到 phase_type（详见 PROJECT_SPEC §5.3 映射表）",
        ))
    return phase_type


def parse_phase_type_cell(value: Any) -> tuple[str, str] | None:
    """解析新格式"阶段类型"列（如 'P4 工业设计'）→ (phase_type, name)。

    - 带前缀（P1-P8 + 名称）→ 拆分；全角 'P１ 需求评估' 经 NFKC 清洗后同样命中
    - 仅前缀（'P4'）→ name 用前缀本身
    - 无前缀纯文本（'工业设计'）→ 查映射表归一化索引，命中 → (type, 原文本)；
      未命中 → None（上层警告 + 跳过该行，决策 2：不猜、不吞）
    """
    if value is None:
        return None
    text = clean_cell(value) if isinstance(value, str) else str(value).strip()
    if not text:
        return None
    m = _PHASE_TYPE_PREFIX_RE.match(text)
    if m and m.group(1):
        name = m.group(2).strip() or m.group(1)
        return m.group(1), name
    # 无前缀：映射表归一化索引兜底
    phase_type = _PHASE_NAME_LOOKUP.get(_normalize_name(text))
    if phase_type is not None:
        return phase_type, text
    return None


def _market_from_sheet(sheet_name: str) -> str:
    """旧格式：根据 sheet 名推断市场。"""
    if "海外" in sheet_name:
        return "海外"
    return "国内"


def detect_format(ws) -> str:
    """按表头识别格式：'new'（14列模板）| 'old'（8列历史）| 'unknown'。"""
    header = ""
    for r in (1, 2):
        for c in range(1, 16):
            v = ws.cell(r, c).value
            if v is not None:
                header += str(v)
    if all(k in header for k in _NEW_FORMAT_KEYS):
        return "new"
    if all(k in header for k in _OLD_FORMAT_KEYS):
        return "old"
    return "unknown"


# ---------- 资源缓存（一次导入内复用） ----------

def _get_or_create_resource(db: Session, cache: dict[str, Resource], name: str) -> Resource:
    """按姓名查/建 Resource，缓存避免重复查询。"""
    if name in cache:
        return cache[name]
    r = db.scalars(select(Resource).where(Resource.name == name)).first()
    if r is None:
        r = Resource(name=name)
        db.add(r)
        db.flush()
    cache[name] = r
    return r


def _special_referenced_resource_ids():
    """专项项目阶段引用的 Resource id 子查询（SPECIAL_PROJECT §五·B）。

    常规导入（全量重置）清理 Resource 时的保留集：专项 assignees 全局复用
    Resource 行，若被常规导入删除会导致专项阶段 assignee 断裂。
    """
    return (
        select(Resource.id)
        .join(phase_assignee, phase_assignee.c.resource_id == Resource.id)
        .join(Phase, Phase.id == phase_assignee.c.phase_id)
        .join(Project, Project.id == Phase.project_id)
        .where(Project.is_special.is_(True))
    )


# ---------- 内存数据结构（解析阶段产出，零 DB 副作用） ----------

@dataclass
class ParsedPhase:
    """解析出的阶段（未落库）。"""
    name: str
    phase_type: str
    sequence: int
    plan_start: date | None = None
    plan_end: date | None = None
    actual_start: date | None = None
    actual_end: date | None = None
    status: str = "未开始"
    progress: int = 0
    remark: str | None = None
    assignees: list[str] = field(default_factory=list)
    # 显式标记：单元格是否有值（合并模式下"未填"不覆盖系统值）
    status_explicit: bool = False
    progress_explicit: bool = False


@dataclass
class ParsedProject:
    """解析出的项目（未落库），phases 内按 sequence 有序。"""
    code: str
    category: str
    name: str
    owner: str
    market: str
    status: str
    plan_start: date | None = None
    plan_end: date | None = None
    remark: str | None = None
    phases: list[ParsedPhase] = field(default_factory=list)
    # 显式标记：项目状态单元格是否有值（合并模式下"未填"不覆盖系统状态）
    status_explicit: bool = False


@dataclass
class ParsedWorkbook:
    """解析结果：项目列表 + 解析报告（errors/warnings/统计）。"""
    projects: list[ParsedProject]
    report: ImportReport


# ---------- 阶段一：解析（纯函数，不碰数据库） ----------

def parse_workbook(file_bytes: bytes, default_category: str = "新需求", special: bool = False) -> ParsedWorkbook:
    """解析 Excel 文件为内存数据结构。

    - 不访问数据库、无任何副作用
    - 返回 ParsedWorkbook：项目列表 + 报告（错误/警告/行数统计）
    - 文件级错误（无法读取/无数据 sheet）时 projects 为空、errors 非空
    - special=True（专项导入，SPECIAL_PROJECT §五·B）：阶段类型列**原样存储**
      （trim 后直接作为 phase_type，不映射 P1-P8、不做兜底/警告；
      旧格式无类型列 → phase_type 留空）
    """
    report = ImportReport()
    projects: list[ParsedProject] = []

    # 解析工作簿
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        report.errors.append(ImportError(
            row=0, sheet="(文件)", field="file",
            message=f"无法读取 Excel 文件: {e}",
        ))
        return ParsedWorkbook(projects, report)

    # 选 sheet：识别出格式（new/old）的数据 sheet；'unknown'（如"填报说明"）跳过
    target_sheets = [(name, detect_format(wb[name])) for name in wb.sheetnames]
    target_sheets = [(n, f) for n, f in target_sheets if f != "unknown"]

    if not target_sheets:
        report.errors.append(ImportError(
            row=0, sheet="(文件)", field="file",
            message=(
                "Excel 中未找到可识别的数据工作表。请使用「项目填报模板.xlsx」"
                "（表头含：项目编号/项目类目/项目名称/项目负责人/市场/阶段类型）"
            ),
        ))
        return ParsedWorkbook(projects, report)

    # 全局项目编号计数器：跨 sheet 连续编号（1,2,3...），避免两表编号冲突
    project_seq = 0

    for sheet_name, fmt in target_sheets:
        ws = wb[sheet_name]
        # 当前项目（阶段挂载用）：遇到项目行时更新
        current_project: ParsedProject | None = None
        # 当前项目在文件内的编号（§2.2 归属校验用；与系统顺序编号不同——
        # 用户表内项目编号常按区块从 1 重排，比对必须用文件内编号）
        current_file_code = ""

        for r in range(_DATA_START_ROW, ws.max_row + 1):
            if fmt == "new":
                row_vals = _read_new_format_row(ws, r)
            else:
                row_vals = _read_old_format_row(ws, r)
            code, category, name, owner, plan_start_raw, plan_end_raw, status = row_vals[:7]
            actual_start_raw, actual_end_raw, phase_type_cell, assignees, progress, remark, market = row_vals[7:]

            report.total_rows += 1
            row_kind = classify_row(code, name, phase_type_cell)

            if row_kind == "skip":
                # §2.7 跳过可见化：空行静默，其余（有内容但无法识别）记警告
                if any(v is not None and str(v).strip() != "" for v in row_vals):
                    report.warnings.append(ImportWarning(
                        row=r, sheet=sheet_name, field="row",
                        message=(
                            f"第{r}行无法识别（编号='{code}' 阶段类型='{phase_type_cell}'），已跳过"
                        ),
                    ))
                continue

            # 文本列统一走 clean_cell（§2.3）。注意数值 0 是有效值，不能 or-丢弃。
            code_str = _clean_str(code)
            name_str = _clean_str(name)
            status_str = _clean_str(status) or "未开始"
            status_explicit = _clean_str(status) != ""

            if row_kind == "project":
                # 项目行
                owner_str = _clean_str(owner)
                plan_start = parse_cell_date(plan_start_raw, r, sheet_name, "plan_start", report)
                plan_end = parse_cell_date(plan_end_raw, r, sheet_name, "plan_end", report)
                market_str = _clean_str(market) or (
                    _market_from_sheet(sheet_name) if fmt == "old" else ""
                )

                # 项目编号：按导入顺序连续编号（1,2,3...），跨 sheet 统一序列，不分市场。
                project_seq += 1
                unique_code = str(project_seq)

                # 新格式优先用列内类目；旧格式沿用 default_category（历史行为）
                cat = _clean_str(category) if fmt == "new" and category is not None else default_category

                current_project = ParsedProject(
                    code=unique_code,
                    category=cat or default_category,
                    name=name_str,
                    owner=owner_str,
                    market=market_str or "国内",
                    status=status_str or "未开始",
                    plan_start=plan_start,
                    plan_end=plan_end,
                    remark=_clean_str(remark) or None,
                    status_explicit=status_explicit,
                )
                projects.append(current_project)
                report.projects_imported += 1
                # 记录文件内编号，供阶段行归属校验（§2.2）
                current_file_code = code_str

            elif row_kind == "phase":
                # 阶段行：归属 current_project（§2.2）
                if current_project is None:
                    report.errors.append(ImportError(
                        row=r, sheet=sheet_name, field="project_id",
                        message=f"阶段'{name_str}'缺少父项目（项目编号 {code}），跳过",
                    ))
                    continue

                # §2.2 归属校验：父号 ≠ 当前项目行（文件内）编号 → 警告 + 仍归属 current
                parent_m = _PHASE_PARENT_RE.match(code_str)
                if parent_m and current_file_code and parent_m.group(1) != current_file_code:
                    report.warnings.append(ImportWarning(
                        row=r, sheet=sheet_name, field="code",
                        message=(
                            f"阶段编号'{code_str}'与当前项目（{current_file_code}）"
                            f"不一致，已归属项目'{current_project.name}'"
                        ),
                    ))

                # 阶段类型与名称：新格式优先解析"阶段类型"列。
                # - 带前缀（'P4 工业设计'）→ 拆 code + name
                # - 纯中文（'工业设计'）→ 映射表归一化索引兜底
                # - 名称列有值时优先用名称列
                # - 旧格式无阶段类型列 → 名称列走 map_phase_type（error 兜底，历史行为）
                # - 专项模式（special=True，§五·B）：阶段类型列**原样存储**——
                #   不映射 P1-P8、不做兜底/警告；旧格式无类型列 → phase_type 留空
                if special:
                    phase_type = _clean_str(phase_type_cell)
                    phase_name = name_str or phase_type
                else:
                    pt = parse_phase_type_cell(phase_type_cell)
                    if pt:
                        phase_type, type_name = pt
                        phase_name = name_str or type_name
                    elif fmt == "new" and phase_type_cell is not None and _clean_str(phase_type_cell):
                        # 新格式：阶段类型列有值但不可解析 → 警告 + 跳过该行（§2.6 决策 2）
                        report.warnings.append(ImportWarning(
                            row=r, sheet=sheet_name, field="phase_type",
                            message=f"阶段类型'{_clean_str(phase_type_cell)}'无法识别，已跳过该行",
                        ))
                        continue
                    else:
                        phase_type = None
                        phase_name = name_str or _clean_str(phase_type_cell)
                    if not phase_type:
                        phase_type = map_phase_type(phase_name, r, sheet_name, report)
                        if phase_type is None:
                            # 旧格式名称兜底失败（新格式无阶段类型列值的路径不会到这）：
                            # 名称列也无法映射 → 不猜、不吞，跳过该行（§2.6 决策 2）
                            report.warnings.append(ImportWarning(
                                row=r, sheet=sheet_name, field="phase_type",
                                message=f"阶段类型'{phase_name}'无法识别，已跳过该行",
                            ))
                            continue

                plan_start = parse_cell_date(plan_start_raw, r, sheet_name, "plan_start", report)
                plan_end = parse_cell_date(plan_end_raw, r, sheet_name, "plan_end", report)
                actual_start = parse_cell_date(actual_start_raw, r, sheet_name, "actual_start", report)
                actual_end = parse_cell_date(actual_end_raw, r, sheet_name, "actual_end", report)

                # sequence：子序号提取（'1-3'/'1.3' → 3）；无编号按追加序号
                seq_match = re.search(r"[-.](\d+)$", code_str)
                sequence = int(seq_match.group(1)) if seq_match else len(current_project.phases) + 1

                # 进度：统一 0-100（§2.5）；显式列优先，无法解析 → 警告 + 按状态推断
                progress_text = _clean_str(progress)
                progress_explicit = progress_text not in ("", "/")
                progress_val: int | None = parse_progress(progress) if progress_explicit else None
                if progress_explicit and progress_val is None:
                    report.warnings.append(ImportWarning(
                        row=r, sheet=sheet_name, field="progress",
                        message=f"无法解析进度 '{progress}'，已按状态推断",
                    ))
                if progress_val is None:
                    progress_explicit = False
                    progress_val = 100 if status_str == "已完成" else (50 if status_str == "进行中" else 0)

                current_project.phases.append(ParsedPhase(
                    name=phase_name,
                    phase_type=phase_type,
                    sequence=sequence,
                    plan_start=plan_start,
                    plan_end=plan_end,
                    actual_start=actual_start,
                    actual_end=actual_end,
                    status=status_str or "未开始",
                    progress=progress_val,
                    remark=_clean_str(remark) or None,
                    assignees=split_persons(assignees),
                    status_explicit=status_explicit,
                    progress_explicit=progress_explicit,
                ))
                report.phases_imported += 1

    return ParsedWorkbook(projects, report)


# ---------- 阶段二：落库 ----------

def import_parsed(db: Session, parsed: ParsedWorkbook) -> ImportReport:
    """把解析结果写入数据库（全量重置：删所有常规 Project/Resource，保留 Template）。

    - 专项项目完全隔离（SPECIAL_PROJECT §五·B）：只删 `is_special=False` 的项目；
      Resource 只删"专项未引用的行"（专项 assignees 全局复用 Resource，不能断裂）；
      报告 warnings 追加"专项项目 N 个不受影响"信息
    - parsed.report 中的 errors/warnings/统计原样保留，补充 resources_created
    - 返回 ImportReport，并存为最近报告供 GET /api/import/report 查询
    """
    report = parsed.report

    # 文件级错误（无法读取/无数据 sheet）：不落库，直接返回
    if not parsed.projects and report.errors:
        _set_last_report(report)
        return report

    # 全量重置（常规域）：删非专项 Project（级联）+ Resource（保留 Template）
    # 先解除 user→resource 引用：PG 外键约束下直接删被引用的 resource 会失败
    special_count = db.query(Project).filter(Project.is_special.is_(True)).count()
    db.execute(update(User).values(resource_id=None).where(User.resource_id.is_not(None)))
    db.query(Project).filter(Project.is_special.is_(False)).delete()
    db.query(Resource).filter(Resource.id.not_in(_special_referenced_resource_ids())).delete()
    db.flush()
    if special_count:
        report.warnings.append(ImportWarning(
            row=0, sheet="(导入)", field="is_special",
            message=f"专项项目 {special_count} 个不受影响（常规导入不重置专项域）",
        ))

    resource_cache: dict[str, Resource] = {}
    resources_created = 0
    # 项目编号续编起点：现有最大纯数字编号 + 1（保留的专项项目可能占用 '1'..'N'，
    # 直接用文件行号会撞唯一约束 project.code——实测 PG 复现）
    code_base = int(_next_project_code(db))

    for i, project in enumerate(parsed.projects):
        proj = Project(
            code=str(code_base + i),
            category=project.category,
            name=project.name,
            owner=project.owner,
            market=project.market,
            status=project.status,
            plan_start=project.plan_start,
            plan_end=project.plan_end,
            remark=project.remark,
        )
        db.add(proj)
        db.flush()
        # 项目负责人也作为资源
        if project.owner:
            before = len(resource_cache)
            _get_or_create_resource(db, resource_cache, project.owner)
            if len(resource_cache) > before:
                resources_created += 1

        for ph in project.phases:
            phase = Phase(
                project_id=proj.id,
                phase_type=ph.phase_type,
                name=ph.name,
                sequence=ph.sequence,
                plan_start=ph.plan_start,
                plan_end=ph.plan_end,
                actual_start=ph.actual_start,
                actual_end=ph.actual_end,
                status=ph.status,
                progress=ph.progress,
                rework_count=0,
                remark=ph.remark,
            )
            db.add(phase)
            db.flush()

            # 拆分负责人，建/匹配 Resource 并关联
            if ph.assignees:
                assignee_resources = []
                for pname in ph.assignees:
                    before = len(resource_cache)
                    res = _get_or_create_resource(db, resource_cache, pname)
                    if len(resource_cache) > before:
                        resources_created += 1
                    assignee_resources.append(res)
                phase.assignees = assignee_resources

    # 收尾：对所有项目的阶段按 sequence 建 FS 串联依赖
    _build_all_project_dependencies(db)

    report.resources_created = resources_created
    db.commit()

    _set_last_report(report)
    return report


def import_excel(db: Session, file_bytes: bytes, default_category: str = "新需求") -> ImportReport:
    """解析 Excel 文件并全量导入（兼容入口 = parse_workbook + import_parsed）。"""
    return import_parsed(db, parse_workbook(file_bytes, default_category))


def import_special(db: Session, parsed: ParsedWorkbook) -> ImportReport:
    """把解析结果写入数据库（专项域全量重置，SPECIAL_PROJECT §五·B）。

    - 删全部 `is_special=true` 项目（阶段/assignee 级联），按文件重建（is_special=True）
    - 常规项目与人员 Resource 完全不受影响（Resource 全局复用）
    - 阶段类型为文件原样存储（parse_workbook special 模式解析）
    - 收尾建专项域 FS 串联链；报告复用 ImportReport 结构，并存为最近报告
    """
    report = parsed.report

    # 文件级错误（无法读取/无数据 sheet）：不落库，直接返回
    if not parsed.projects and report.errors:
        _set_last_report(report)
        return report

    # 全量重置专项域：删全部 is_special=true 项目（DB 级联删阶段/assignee/rework_log）
    db.query(Project).filter(Project.is_special.is_(True)).delete()
    db.flush()

    resource_cache: dict[str, Resource] = {}
    resources_created = 0
    # 项目编号续编起点：现有最大纯数字编号 + 1（常规项目可能占用 '1'..'N'，
    # 直接用文件行号会撞唯一约束 project.code——与 import_parsed 同口径）
    code_base = int(_next_project_code(db))

    for i, project in enumerate(parsed.projects):
        proj = Project(
            code=str(code_base + i),
            category=project.category,
            name=project.name,
            owner=project.owner,
            market=project.market,
            status=project.status,
            plan_start=project.plan_start,
            plan_end=project.plan_end,
            remark=project.remark,
            is_special=True,
        )
        db.add(proj)
        db.flush()
        # 项目负责人也作为资源
        if project.owner:
            before = len(resource_cache)
            _get_or_create_resource(db, resource_cache, project.owner)
            if len(resource_cache) > before:
                resources_created += 1

        for ph in project.phases:
            phase = Phase(
                project_id=proj.id,
                phase_type=ph.phase_type,
                name=ph.name,
                sequence=ph.sequence,
                plan_start=ph.plan_start,
                plan_end=ph.plan_end,
                actual_start=ph.actual_start,
                actual_end=ph.actual_end,
                status=ph.status,
                progress=ph.progress,
                rework_count=0,
                remark=ph.remark,
            )
            db.add(phase)
            db.flush()

            # 拆分负责人，建/匹配 Resource 并关联（人员全局复用）
            if ph.assignees:
                assignee_resources = []
                for pname in ph.assignees:
                    before = len(resource_cache)
                    res = _get_or_create_resource(db, resource_cache, pname)
                    if len(resource_cache) > before:
                        resources_created += 1
                    assignee_resources.append(res)
                phase.assignees = assignee_resources

    # 收尾：专项域项目按 sequence 建 FS 串联依赖
    _build_all_project_dependencies(db, special_only=True)

    report.resources_created = resources_created
    db.commit()

    _set_last_report(report)
    return report


# ---------- 阶段三：增量合并导入（默认模式） ----------

# 纯数字编号正则（合并模式新项目自动编号用）
_CODE_INT_RE = re.compile(r"^\d+$")


def _next_project_code(db: Session) -> str:
    """生成下一个项目编号：现有纯数字编号最大值 + 1（无则从 1 开始）。"""
    codes = db.scalars(select(Project.code)).all()
    nums = [int(c) for c in codes if c and _CODE_INT_RE.match(c)]
    return str(max(nums) + 1) if nums else "1"


# 阶段类型自然顺序（P1 最前 ... P8 最后），用于新增阶段的插入排序
_PHASE_TYPE_ORDER = {f"P{i}": i for i in range(1, 9)}


def _phase_type_rank(phase_type: str) -> int:
    return _PHASE_TYPE_ORDER.get(phase_type, 9)


def _merge_project(
    db: Session,
    project: Project,
    parsed_p: ParsedProject,
    resource_cache: dict[str, Resource],
    counters: dict[str, int],
    pending_links: list[tuple[str, str]],
) -> None:
    """合并同名项目：项目字段非空覆盖 + 阶段按类型匹配更新/自然序插入新增。

    counters: {"phases_created", "phases_updated"} 统计累加
    pending_links: 收集"新增阶段待关联依赖"（项目名, 阶段名）
    """
    # ---------- 项目字段：非空覆盖 ----------
    if parsed_p.owner:
        project.owner = parsed_p.owner
    if parsed_p.market:
        project.market = parsed_p.market
    if parsed_p.category:
        project.category = parsed_p.category
    if parsed_p.plan_start:
        project.plan_start = parsed_p.plan_start
    if parsed_p.plan_end:
        project.plan_end = parsed_p.plan_end
    if parsed_p.remark:
        project.remark = parsed_p.remark
    if parsed_p.status_explicit:
        project.status = parsed_p.status

    # ---------- 阶段合并 ----------
    existing = sorted(project.phases, key=lambda p: p.sequence)
    # 类型 → 阶段映射（同类型多个时只匹配第一个，其余文件同类型按新增处理）
    by_type: dict[str, Phase] = {}
    for ph in existing:
        by_type.setdefault(ph.phase_type, ph)

    for pp in parsed_p.phases:
        ph = by_type.get(pp.phase_type)
        if ph is not None:
            _merge_phase(db, ph, pp, resource_cache)
            counters["phases_updated"] += 1
        else:
            new_phase = _insert_phase(db, project.id, existing, pp, resource_cache)
            counters["phases_created"] += 1
            pending_links.append((project.name, pp.name))
            existing = sorted(project.phases, key=lambda p: p.sequence)
            by_type.setdefault(pp.phase_type, new_phase)

    # 负责人：项目负责人也作为资源（只增不删）
    if project.owner:
        _get_or_create_resource(db, resource_cache, project.owner)


def _merge_phase(db: Session, phase: Phase, pp: ParsedPhase, resource_cache: dict[str, Resource]) -> None:
    """更新已有阶段：文件非空字段覆盖；未填保留系统值。"""
    if pp.name:
        phase.name = pp.name
    if pp.plan_start:
        phase.plan_start = pp.plan_start
    if pp.plan_end:
        phase.plan_end = pp.plan_end
    if pp.actual_start:
        phase.actual_start = pp.actual_start
    if pp.actual_end:
        phase.actual_end = pp.actual_end
    if pp.status_explicit:
        phase.status = pp.status
    if pp.progress_explicit:
        phase.progress = pp.progress
    if pp.remark:
        phase.remark = pp.remark
    if pp.assignees:
        # 文件有值 → 覆盖负责人列表（resource 本身只增不删）
        assignees = []
        for pname in pp.assignees:
            assignees.append(_get_or_create_resource(db, resource_cache, pname))
        phase.assignees = assignees


def _insert_phase(
    db: Session,
    project_id: int,
    existing: list[Phase],
    pp: ParsedPhase,
    resource_cache: dict[str, Resource],
) -> Phase:
    """按阶段类型自然顺序插入新阶段（P1 最前 ... P8 最后），后续阶段序号顺延。返回新阶段。"""
    # 找插入点：第一个类型顺序大于新阶段的现有阶段
    pos = 0
    for ph in existing:
        if _phase_type_rank(ph.phase_type) > _phase_type_rank(pp.phase_type):
            break
        pos += 1

    # 插入点及之后的阶段序号顺延
    for ph in existing[pos:]:
        ph.sequence += 1

    if pos == 0:
        new_seq = 1
    else:
        new_seq = existing[pos - 1].sequence + 1

    phase = Phase(
        project_id=project_id,
        phase_type=pp.phase_type,
        name=pp.name,
        sequence=new_seq,
        plan_start=pp.plan_start,
        plan_end=pp.plan_end,
        actual_start=pp.actual_start,
        actual_end=pp.actual_end,
        status=pp.status,
        progress=pp.progress,
        rework_count=0,
        remark=pp.remark,
    )
    db.add(phase)
    db.flush()
    if pp.assignees:
        assignee_resources = []
        for pname in pp.assignees:
            assignee_resources.append(_get_or_create_resource(db, resource_cache, pname))
        phase.assignees = assignee_resources
    return phase


def import_merged(db: Session, parsed: ParsedWorkbook) -> ImportReport:
    """增量合并导入（默认模式）：新增 + 更新 + 保留，**不删除任何现有数据**。

    - 同名项目 → 项目字段非空覆盖 + 阶段按类型匹配更新/自然序插入新增
    - 新项目 → 创建（含 FS 串联依赖）
    - 已有依赖完全不动；同名项目的新增阶段不自动建依赖（报告提示待关联）
    """
    report = parsed.report

    # 文件级错误（无法读取/无数据 sheet）：不落库，直接返回
    if not parsed.projects and report.errors:
        _set_last_report(report)
        return report

    resource_cache: dict[str, Resource] = {}
    resources_created = 0
    counters = {"phases_created": 0, "phases_updated": 0}
    pending_links: list[tuple[str, str]] = []
    created_projects: list[Project] = []  # 新建项目（收尾建 FS 链）

    for parsed_p in parsed.projects:
        # 比较键归一化（评审处置 #2）：全角/半角书写差异不影响同名命中
        # 专项项目不参与合并匹配（§五·B）：与专项同名 → 视为新建常规项目
        existing = db.scalars(
            select(Project).where(
                Project.name == parsed_p.name,
                Project.is_special.is_(False),
            )
        ).first()
        if existing is None:
            # 归一化后重查（项目名 NFKC 不同但语义相同 → 合并而非新增）
            existing = next(
                (p for p in db.scalars(
                    select(Project).where(Project.is_special.is_(False))
                ).all()
                 if _normalize_name(p.name or "") == _normalize_name(parsed_p.name)),
                None,
            )
        if existing:
            _merge_project(db, existing, parsed_p, resource_cache, counters, pending_links)
            report.projects_updated += 1
        else:
            # 新建项目：项目编号沿用系统自动编号（解析编号仅作行归类用，避免与现有项目冲突）
            proj = Project(
                code=_next_project_code(db),
                category=parsed_p.category,
                name=parsed_p.name,
                owner=parsed_p.owner,
                market=parsed_p.market,
                status=parsed_p.status,
                plan_start=parsed_p.plan_start,
                plan_end=parsed_p.plan_end,
                remark=parsed_p.remark,
            )
            db.add(proj)
            db.flush()
            if proj.owner:
                before = len(resource_cache)
                _get_or_create_resource(db, resource_cache, proj.owner)
                if len(resource_cache) > before:
                    resources_created += 1

            for pp in parsed_p.phases:
                phase = Phase(
                    project_id=proj.id,
                    phase_type=pp.phase_type,
                    name=pp.name,
                    sequence=pp.sequence,
                    plan_start=pp.plan_start,
                    plan_end=pp.plan_end,
                    actual_start=pp.actual_start,
                    actual_end=pp.actual_end,
                    status=pp.status,
                    progress=pp.progress,
                    rework_count=0,
                    remark=pp.remark,
                )
                db.add(phase)
                db.flush()
                if pp.assignees:
                    assignee_resources = []
                    for pname in pp.assignees:
                        before = len(resource_cache)
                        res = _get_or_create_resource(db, resource_cache, pname)
                        if len(resource_cache) > before:
                            resources_created += 1
                        assignee_resources.append(res)
                    phase.assignees = assignee_resources
            created_projects.append(proj)
            report.projects_created += 1
            counters["phases_created"] += len(parsed_p.phases)

    # 新建项目：按 sequence 建 FS 串联链（同名项目已有依赖不动）
    for proj in created_projects:
        phases = list(db.scalars(
            select(Phase).where(Phase.project_id == proj.id).order_by(Phase.sequence)
        ))
        _build_sequence_dependencies(db, phases)

    report.phases_created = counters["phases_created"]
    report.phases_updated = counters["phases_updated"]
    report.resources_created = resources_created
    report.pending_link_phases = [
        PendingLinkPhase(project_name=pn, phase_name=phn) for pn, phn in pending_links
    ]
    db.commit()

    _set_last_report(report)
    return report


# ---------- 导入前差异报告（预览） ----------

def build_preview(db: Session, parsed: ParsedWorkbook, special: bool = False) -> ImportPreview:
    """基于解析结果 + 当前库生成差异报告（只读统计，无任何副作用）。

    包含两类信息：
    - 全量替换视角：existing（将被清空）/ incoming / match（同名对比）
    - 增量合并视角：created/updated/kept 明细 + 新增阶段待关联依赖提示

    special=True（专项导入预览）：existing/同名对比/合并明细全部按专项域口径
    （常规项目不计入；Resource 不清零——专项导入不删 Resource，人员全局复用）。
    """
    if special:
        # 专项域口径（SPECIAL_PROJECT §五·B 专项预览）
        existing_projects = db.query(Project).filter(Project.is_special.is_(True)).count()
        existing_phases = (
            db.query(Phase).join(Project).filter(Project.is_special.is_(True)).count()
        )
        existing_resources = 0  # 专项导入全量重置不删 Resource（人员全局复用）
        existing_query = select(Project).where(Project.is_special.is_(True))
    else:
        # 常规域口径：专项项目不计入（将被清空）/同名对比/合并明细
        existing_projects = db.query(Project).filter(Project.is_special.is_(False)).count()
        existing_phases = (
            db.query(Phase).join(Project).filter(Project.is_special.is_(False)).count()
        )
        # 常规导入实际只删"专项未引用的 Resource 行"——与 import_parsed 清理口径一致
        existing_resources = (
            db.query(Resource).filter(Resource.id.not_in(_special_referenced_resource_ids())).count()
        )
        existing_query = select(Project).where(Project.is_special.is_(False))

    incoming_projects = len(parsed.projects)
    incoming_phases = sum(len(p.phases) for p in parsed.projects)

    # 同名项目对比（防传错文件；比较键 NFKC 归一化，评审处置 #2）
    existing_names_raw = set(db.scalars(existing_query.with_only_columns(Project.name)).all())
    existing_names = {_normalize_name(n or "") for n in existing_names_raw}
    incoming_names = {_normalize_name(p.name) for p in parsed.projects if p.name}
    matched = len(incoming_names & existing_names)
    new = len(incoming_names - existing_names)
    missing = len(existing_names - incoming_names)

    def _preview(p: ParsedProject) -> PreviewProject:
        return PreviewProject(
            name=p.name,
            market=p.market,
            category=p.category,
            phases=len(p.phases),
        )

    # 项目概览（前 20 个）
    projects_preview = [_preview(p) for p in parsed.projects[:20]]

    # 增量合并明细（key 为归一化项目名，与 import_merged 命中口径一致）
    existing_proj_map = {
        _normalize_name(p.name): p for p in db.scalars(existing_query).all()
    }
    created_projects = [
        _preview(p) for p in parsed.projects
        if _normalize_name(p.name) not in existing_proj_map
    ]
    updated_projects = [
        _preview(p) for p in parsed.projects
        if _normalize_name(p.name) in existing_proj_map
    ]

    # 阶段统计与待关联提示（基于现有同名项目的阶段类型）
    phases_created = 0
    phases_updated = 0
    pending: list[PendingLinkPhase] = []
    for p in parsed.projects:
        ep = existing_proj_map.get(_normalize_name(p.name))
        if ep is None:
            phases_created += len(p.phases)  # 新项目全部阶段为新增
            continue
        existing_types = {ph.phase_type for ph in ep.phases}
        for pp in p.phases:
            if pp.phase_type in existing_types:
                phases_updated += 1
            else:
                phases_created += 1
                pending.append(PendingLinkPhase(project_name=p.name, phase_name=pp.name))

    return ImportPreview(
        existing=ImportExistingCounts(
            projects=existing_projects,
            phases=existing_phases,
            resources=existing_resources,
        ),
        incoming=ImportIncomingCounts(
            projects=incoming_projects,
            phases=incoming_phases,
        ),
        match=ProjectMatchInfo(matched=matched, new=new, missing=missing),
        errors=list(parsed.report.errors),
        warnings=list(parsed.report.warnings),
        projects_preview=projects_preview,
        created_projects=created_projects,
        updated_projects=updated_projects,
        kept_count=missing,
        phases_created=phases_created,
        phases_updated=phases_updated,
        pending_link_phases=pending,
    )


def _read_new_format_row(ws, r: int) -> list[Any]:
    """读取 14 列新格式一行，返回固定顺序的值列表。

    返回: [code, category, name, owner, plan_start, plan_end, status,
           actual_start, actual_end, phase_type, assignees, progress, remark, market]
    """
    return [
        ws.cell(r, _NEW_COL_CODE).value,
        ws.cell(r, _NEW_COL_CATEGORY).value,
        ws.cell(r, _NEW_COL_NAME).value,
        ws.cell(r, _NEW_COL_OWNER).value,
        ws.cell(r, _NEW_COL_PLAN_START).value,
        ws.cell(r, _NEW_COL_PLAN_END).value,
        ws.cell(r, _NEW_COL_STATUS).value,
        ws.cell(r, _NEW_COL_ACTUAL_START).value,
        ws.cell(r, _NEW_COL_ACTUAL_END).value,
        ws.cell(r, _NEW_COL_PHASE_TYPE).value,
        ws.cell(r, _NEW_COL_ASSIGNEES).value,
        ws.cell(r, _NEW_COL_PROGRESS).value,
        ws.cell(r, _NEW_COL_REMARK).value,
        ws.cell(r, _NEW_COL_MARKET).value,
    ]


def _read_old_format_row(ws, r: int) -> list[Any]:
    """读取 8 列旧格式一行，返回与新格式相同的顺序。

    返回: [code, category, name, owner, plan_start, plan_end, status,
           actual_start, actual_end, phase_type, assignees, progress, remark, market]
    - 旧格式无 实际日期/阶段类型/进度/备注 列 → None
    - 阶段行负责人 = 第 4 列（负责人列）
    - 交接人（第 8 列）读取后丢弃（字段已废弃）
    """
    return [
        ws.cell(r, _OLD_COL_CODE).value,
        ws.cell(r, _OLD_COL_CATEGORY).value,
        ws.cell(r, _OLD_COL_NAME).value,
        ws.cell(r, _OLD_COL_OWNER).value,
        ws.cell(r, _OLD_COL_PLAN_START).value,
        ws.cell(r, _OLD_COL_PLAN_END).value,
        ws.cell(r, _OLD_COL_STATUS).value,
        None,  # actual_start
        None,  # actual_end
        None,  # phase_type 列（旧格式无，走名称映射）
        ws.cell(r, _OLD_COL_OWNER).value,  # assignees = 负责人列
        None,  # progress
        None,  # remark
        None,  # market（旧格式按 sheet 名推断）
    ]


def _build_sequence_dependencies(db: Session, phases: list[Phase]) -> None:
    """按 sequence 升序，相邻阶段建 FS 依赖（phase[i] → phase[i+1]）。"""
    if len(phases) < 2:
        return
    ordered = sorted(phases, key=lambda p: p.sequence)
    for i in range(len(ordered) - 1):
        frm, to = ordered[i], ordered[i + 1]
        if frm.id == to.id:
            continue
        # 避免重复
        exists = db.scalars(
            select(Dependency).where(
                Dependency.from_phase_id == frm.id,
                Dependency.to_phase_id == to.id,
            )
        ).first()
        if exists:
            continue
        db.add(Dependency(from_phase_id=frm.id, to_phase_id=to.id, type="FS", lag_days=0))


def _build_all_project_dependencies(db: Session, special_only: bool = False) -> None:
    """对所有项目（或仅专项项目）的阶段按 sequence 建 FS 串联依赖。

    导入流程的每个项目阶段建好后调用。
    special_only=True：仅专项域项目（专项导入收尾用，不触碰常规项目已有依赖）。
    """
    q = select(Project)
    if special_only:
        q = q.where(Project.is_special.is_(True))
    projects = list(db.scalars(q))
    for project in projects:
        phases = list(db.scalars(
            select(Phase).where(Phase.project_id == project.id).order_by(Phase.sequence)
        ))
        _build_sequence_dependencies(db, phases)
