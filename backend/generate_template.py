"""生成项目经理填报 Excel 模板（与 docs/项目填报模板.xlsx 格式一致，14 列）。

产出文件：docs/项目填报模板.xlsx
- Sheet1「填报说明」：字段说明 + 数据验证规则
- Sheet2「项目填报」：项目填报区（14 列，无交接人，市场为区域级）

样式与数据验证复用 app.services.excel_format（与导出文件保持完全一致）。
"""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from app.services.excel_format import HEADERS, style_sheet

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "docs" / "项目填报模板.xlsx"

# 样式（填报说明 sheet 用）
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="001529")
NOTE_FONT = Font(name="微软雅黑", size=9, color="666666")
BODY_FONT = Font(name="微软雅黑", size=10)


def _create_instruction_sheet(wb):
    """创建填报说明 Sheet。"""
    ws = wb.active
    ws.title = "填报说明"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 50

    title = ws.cell(row=2, column=2, value="📋 填报说明")
    title.font = TITLE_FONT

    instructions = [
        ("填写规则", ""),
        ("", "1. 项目行：项目编号填纯数字（如 1、2、3），其余列按实际情况填写。"),
        ("", "2. 阶段行：项目编号填「项目号-序号」（如 1-1、2-3）。阶段类型为必填下拉。"),
        ("", "3. 空行（项目编号和项目名称均为空）会被自动跳过，可用于分隔不同项目。"),
        ("", "4. 市场列按销售区域选择（拉美区/西欧区/东欧区/中东区/亚太区/土耳其区/非洲区/北美区/OEM业务部）。"),
        ("", "5. 项目编号仅用于行归类（项目行/阶段行）、阶段归属与排序，导入后系统自动重新编号。"),
        ("", ""),
        ("阶段依赖关系", ""),
        ("", "导入后，系统会自动按阶段序号生成 FS（完成-开始）依赖链："),
        ("", "　例如项目 1 有阶段 1-1→1-2→1-3→1-4，导入后自动串联依赖。"),
        ("", "如需更复杂的依赖类型（SS 并行 / 跨阶段），请在系统甘特图中手动调整。"),
        ("", ""),
        ("字段说明", ""),
        ("项目编号", "项目行纯数字，阶段行数字-数字（如 1-1）。仅用于行归类/阶段归属/排序，导入后系统自动重新编号。"),
        ("项目类目", "必填。下拉选择：新需求 / 量产 / 定制 / 改造。"),
        ("项目名称", "必填。项目的完整名称。"),
        ("项目负责人", "必填。负责该项目的项目经理姓名。"),
        ("市场", "必填。下拉选择销售区域（拉美区/西欧区/东欧区/中东区/亚太区/土耳其区/非洲区/北美区/OEM业务部）。"),
        ("阶段类型", "阶段行必填。下拉选择：需求评估 / 配置评估 / 模块选型 / 工业设计 / 结构设计 / 线缆设计 / 样机打样 / 线缆打样 / 联调测试 / 交付。"),
        ("计划开始/结束", "日期格式 YYYY-MM-DD（如 2026-07-01）。"),
        ("实际开始/结束", "日期格式 YYYY-MM-DD。项目行不填，阶段行按实际情况。"),
        ("阶段负责人", "多人用空格或顿号分隔（如「张三 李四」）。"),
        ("阶段状态", "阶段行，下拉选择：未开始 / 进行中 / 已完成 / 延期 / 搁置。"),
        ("阶段进度", "阶段行，填 0-100 的整数。"),
        ("备注", "可选。任何补充说明。"),
        ("", ""),
        ("导入方式", ""),
        ("", "在平台「项目列表」页面点击「导入 Excel」，选择本文件即可。"),
        ("", "导入前会自动清空现有数据（模板不受影响），支持重复导入。"),
        ("", f"模板生成日期：{date.today().isoformat()}"),
    ]

    for i, (field, desc) in enumerate(instructions):
        r = i + 4
        cell_b = ws.cell(row=r, column=2, value=field)
        cell_c = ws.cell(row=r, column=3, value=desc)
        cell_b.font = Font(name="微软雅黑", size=10, bold=bool(field and desc))
        cell_c.font = NOTE_FONT if not field else Font(name="微软雅黑", size=10)


def _create_data_sheet(wb, title, sample_rows=3):
    """创建数据填报 Sheet（样式与导出文件一致，来自 excel_format）。"""
    ws = wb.create_sheet(title=title)
    # 标题
    ws.append(["智能终端研发项目管理平台 — 项目填报模板"])
    # 表头
    ws.append(HEADERS)
    # 样式 + 数据验证（冻结 A3，数据从第 3 行起）
    style_sheet(ws)

    # 提示注释
    ws.cell(row=3, column=1).comment = openpyxl.comments.Comment(
        "项目行填纯数字（如 1、2）\n阶段行填 项目号-序号（如 1-1）\n空行自动跳过",
        "系统提示"
    )

    # 样例数据
    if sample_rows:
        samples = [
            # 项目 1：招标研发 - col: 编号,类目,名称,负责人,市场,阶段类型,计划开始,计划结束,实际开始,实际结束,阶段负责人,阶段状态,进度,备注
            [1, "新需求", "示例：XX银行自助终端TCM-001", "张三", "拉美区", "", "2026-07-01", "2026-09-30", "", "", "", "", "", "可描述项目内容、背景、风险等"],
            ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-05", "2026-07-01", "2026-07-05", "李四", "已完成", 100, ""],
            ["1-2", "", "", "", "", "结构设计", "2026-07-06", "2026-07-20", "2026-07-06", "", "李四 王五", "进行中", 60, ""],
            ["1-3", "", "", "", "", "联调测试", "2026-07-21", "2026-09-30", "", "", "王五", "未开始", 0, "待启动"],
            # 空行
            ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            # 项目 2：量产 - 海外
            [2, "量产", "示例：海外侧柜量产RS-028", "张三", "中东区", "", "2026-08-01", "2026-12-31", "", "", "", "", "", ""],
            ["2-1", "", "", "", "", "结构设计", "2026-08-01", "2026-08-30", "", "", "钱七", "未开始", 0, ""],
        ]
        for i, row in enumerate(samples):
            r = i + 3
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = BODY_FONT


def main():
    wb = openpyxl.Workbook()

    _create_instruction_sheet(wb)
    _create_data_sheet(wb, "项目填报")

    # 备份旧模板后覆盖
    if OUTPUT.exists():
        bak = OUTPUT.with_suffix(".xlsx.bak")
        shutil.copy2(OUTPUT, bak)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"✅ 模板已生成：{OUTPUT}")
    print(f"   Sheet1「填报说明」— 字段说明 + 导入指引")
    print(f"   Sheet2「项目填报」— 200 行填报区 + 7 项下拉/数值验证")
    print(f"   旧模板已备份：{OUTPUT.with_suffix('.xlsx.bak')}")


if __name__ == "__main__":
    main()
