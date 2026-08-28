"""Excel 导入解析兼容性专项测试（docs/EXCEL_PARSE_COMPAT.md §四，14 用例）。

覆盖：
- §2.1 行判定（阶段类型优先 / 点分编号 / 空编号 / 纯数字编号+阶段类型）
- §2.2 阶段归属（父号不一致警告 / 缺父项目错误）
- §2.3 clean_cell 全角清洗
- §2.4 日期多格式链（点分 / 中文 / 短格式 / 全角 / 8 位数字）
- §2.5 进度统一 0-100（评审处置 #3：输出 0-100，注意 §四表内"0.5"为笔误）
- §2.6 阶段类型不可识别 → 警告 + 跳过
- §2.7 跳过可见化
- 评审处置 #2：合并导入项目名比较键归一化

用例 12 按评审处置 #8 fixture 化：仓库内置同构匿名 fixture；
真实表（项目填报总表202608 (1).xlsx）存在时执行，缺失则 skip。
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.models import Dependency, Phase, Project
from app.schemas.import_report import ImportReport
from app.services.excel_importer import (
    classify_row,
    clean_cell,
    import_excel,
    parse_cell_date,
    parse_phase_type_cell,
    parse_progress,
    parse_workbook,
)

_REAL_EXCEL = Path(r"C:\Users\1\Downloads\项目填报总表202608 (1).xlsx")

_HEADERS = ["项目编号", "项目类目", "项目名称", "项目负责人", "市场", "阶段类型",
            "计划开始", "计划结束", "实际开始", "实际结束",
            "阶段负责人", "阶段状态", "阶段进度", "备注"]


def _empty_report() -> ImportReport:
    return ImportReport()


def _make_workbook(rows: list[list]) -> bytes:
    """按 14 列新格式构造 Excel。每行: [编号, 类目, 名称, 负责人, 市场, 阶段类型,
    计划开始, 计划结束, 实际开始, 实际结束, 阶段负责人, 阶段状态, 阶段进度, 备注]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目填报"
    for c, h in enumerate(_HEADERS, 1):
        ws.cell(1, c, h)
    for i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(3 + i, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- §2.3 clean_cell ----------

def test_clean_cell_fullwidth_and_invisible():
    """§2.3 清洗：全角→半角、不可见字符去除、空白压缩。"""
    assert clean_cell("１－１") == "1-1"
    assert clean_cell("２０２６．６．２９") == "2026.6.29"
    assert clean_cell("ＡＢＣ") == "ABC"
    assert clean_cell("  x\u3000") == "x"            # 全角空格 trim
    assert clean_cell("a\u200bb\ufeffc") == "abc"    # 零宽 + BOM
    assert clean_cell("P1  需求评估") == "P1 需求评估"  # 连续空白压缩
    assert clean_cell(None) is None                  # 非字符串原样
    assert clean_cell(50) == 50


# ---------- §2.1 行判定 ----------

def test_classify_row_phase_type_priority():
    """① 阶段类型可解析 → 阶段行（评审处置 #1：编号格式任意，含纯数字）。"""
    assert classify_row("1.1", None, "P1 需求评估") == "phase"
    assert classify_row("1", None, "P1 需求评估") == "phase"    # 用例 13：纯数字编号
    assert classify_row(None, None, "工业设计") == "phase"        # 无编号（用例 2）
    assert classify_row(None, None, "P１ 需求评估") == "phase"    # 全角 P
    assert classify_row("1", None, "自定义阶段") == "skip"        # 类型不可识别不算 ①


def test_classify_row_project_and_phase_codes():
    """②③ 编号判定：纯数字+名称 → 项目；1-1 / 1.1 → 阶段；其他 → skip。"""
    assert classify_row("1", "项目A", None) == "project"
    assert classify_row("1-1", None, None) == "phase"
    assert classify_row("1.1", None, None) == "phase"      # 用例 1：点分编号
    assert classify_row("１－１", None, None) == "phase"    # 全角编号（NFKC）
    assert classify_row(None, None, None) == "skip"        # 空行
    assert classify_row("abc", None, None) == "skip"       # 用例 9：无法识别


def test_classify_row_pure_digit_code_without_name_not_project():
    """纯数字编号但名称空 → 不是项目行（防误建项目）。"""
    assert classify_row("1", "", None) == "skip"


# ---------- §2.4 日期链 ----------

def test_parse_cell_date_dotted():
    """用例 4：'2026.6.29' → 2026-06-29，无警告。"""
    report = _empty_report()
    assert parse_cell_date("2026.6.29", 1, "s", "plan_start", report) == date(2026, 6, 29)
    assert report.warnings == []


def test_parse_cell_date_chinese():
    """用例 5：'2026年6月29日' → 解析成功。"""
    report = _empty_report()
    assert parse_cell_date("2026年6月29日", 1, "s", "plan_start", report) == date(2026, 6, 29)
    assert report.warnings == []


def test_parse_cell_date_short_format():
    """用例 6：'6.29' → 当年 06-29 + 警告（决策 A）。"""
    report = _empty_report()
    d = parse_cell_date("6.29", 1, "s", "plan_start", report)
    assert d == date(date.today().year, 6, 29)
    assert len(report.warnings) == 1
    assert "缺少年份" in report.warnings[0].message


def test_parse_cell_date_fullwidth():
    """用例 7：全角日期 '２０２６．６．２９' 清洗后正常解析。"""
    report = _empty_report()
    assert parse_cell_date("２０２６．６．２９", 1, "s", "plan_start", report) == date(2026, 6, 29)
    assert report.warnings == []


def test_parse_cell_date_digits8_and_serial_bounds():
    """8 位纯数字 YYYYMMDD / 7 位 YYYYMDD；序列号越界（评审处置 #5）落文本链。"""
    report = _empty_report()
    assert parse_cell_date("20260629", 1, "s", "f", report) == date(2026, 6, 29)
    assert parse_cell_date("2026629", 1, "s", "f", report) == date(2026, 6, 29)
    # 越界序列号：不按序列号解析 → 警告（而非 9999 万年后的荒谬日期）
    report2 = _empty_report()
    assert parse_cell_date(9_999_999, 1, "s", "f", report2) is None
    assert len(report2.warnings) == 1
    # 范围内序列号仍正常（serial 由目标日期反推，自洽验证）
    serial = (date(2026, 6, 29) - date(1899, 12, 30)).days
    report3 = _empty_report()
    assert parse_cell_date(serial, 1, "s", "f", report3) == date(2026, 6, 29)
    assert report3.warnings == []


def test_parse_cell_date_numeric_8_7_digits():
    """评审处置 #5 补全：数值型 8/7 位 YYYYMMDD/YYYYMDD（Excel 存为数字的日期）。"""
    report = _empty_report()
    assert parse_cell_date(20260629, 1, "s", "f", report) == date(2026, 6, 29)
    assert parse_cell_date(2026629, 1, "s", "f", report) == date(2026, 6, 29)
    assert report.warnings == []
    # 7 位非日期数值（月 0 非法）在序列号范围内 → 仍按序列号解析（语义不变）
    serial2 = (date(2026, 6, 29) - date(1899, 12, 30)).days
    report_serial = _empty_report()
    assert parse_cell_date(serial2, 1, "s", "f", report_serial) == date(2026, 6, 29)
    assert report_serial.warnings == []


def test_parse_cell_date_numeric_invalid_month_or_day():
    """数值型 8 位但月/日非法（20261399 → 月 13）：警告 + None，不猜日期。"""
    report = _empty_report()
    assert parse_cell_date(20261399, 1, "s", "f", report) is None
    assert len(report.warnings) == 1
    assert "20261399" in report.warnings[0].message
    # 越界 + 非法：同样警告（9_999_999 → 月 99 非法，不按序列号）
    report2 = _empty_report()
    assert parse_cell_date(9_999_999, 1, "s", "f", report2) is None
    assert len(report2.warnings) == 1



def test_parse_cell_date_iso_still_works():
    """回归：ISO 'YYYY-MM-DD' 与斜杠格式仍解析。"""
    report = _empty_report()
    assert parse_cell_date("2026-06-29", 1, "s", "f", report) == date(2026, 6, 29)
    assert parse_cell_date("2026/6/29", 1, "s", "f", report) == date(2026, 6, 29)
    assert report.warnings == []


# ---------- §2.5 进度 ----------

def test_parse_progress_unified_0_100():
    """用例 8：50% / 50％ / 50 % / 0.5 / 50 → 统一 0-100（评审处置 #3）。"""
    assert parse_progress("50%") == 50
    assert parse_progress("50％") == 50      # 全角％（NFKC → %）
    assert parse_progress("50 %") == 50
    assert parse_progress(0.5) == 50
    assert parse_progress("0.5") == 50
    assert parse_progress("50") == 50
    assert parse_progress(50) == 50
    assert parse_progress(100) == 100
    assert parse_progress(0) == 0
    assert parse_progress(150) == 100        # 越界夹取
    assert parse_progress("abc") is None     # 无法解析 → None
    assert parse_progress("/") is None


# ---------- §2.6 阶段类型 ----------

def test_parse_phase_type_cell_variants():
    """P 前缀（含子编号 P71/P72）/ 全角 P / 纯中文名（映射表归一化索引）。"""
    assert parse_phase_type_cell("P1 需求评估") == ("P1", "需求评估")
    assert parse_phase_type_cell("P１ 需求评估") == ("P1", "需求评估")   # 用例 7
    assert parse_phase_type_cell("P4") == ("P4", "P4")
    assert parse_phase_type_cell("P71 样机打样") == ("P71", "样机打样")  # PHASE_TYPES_V2 子编号
    assert parse_phase_type_cell("P72线缆打样") == ("P72", "线缆打样")
    assert parse_phase_type_cell("工业设计") == ("P4", "工业设计")
    # 评审处置 #4：全角括号键经 NFKC 索引可匹配
    assert parse_phase_type_cell("样机打样(1台)") == ("P71", "样机打样(1台)")
    assert parse_phase_type_cell("自定义阶段") is None


# ---------- 端到端：行判定 / 归属 / 可见化 ----------

def test_dotted_code_phase_attached_with_dependency(client, db_session):
    """用例 1：项目行 + 点分编号 1.1/1.2 阶段行 → 归属项目1，依赖按序生成。"""
    data = _make_workbook([
        [1, "新需求", "点分项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1.1", "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "已完成", 100, ""],
        ["1.2", "", "", "", "", "P5 结构设计", "2026-07-11", "2026-07-30", "", "", "王五", "进行中", 60, ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 1
    assert report.phases_imported == 2
    assert report.errors == []
    proj = db_session.query(Project).one()
    phases = db_session.query(Phase).order_by(Phase.sequence).all()
    assert all(ph.project_id == proj.id for ph in phases)
    assert [ph.sequence for ph in phases] == [1, 2]   # 点分子序号 1.1→1、1.2→2
    deps = db_session.query(Dependency).all()
    assert len(deps) == 1 and deps[0].type == "FS"    # 按序 FS 串联


def test_empty_code_phase_belongs_to_previous_project(client, db_session):
    """用例 2：无编号阶段行（有阶段类型）→ 归属上一项目。"""
    data = _make_workbook([
        [1, "新需求", "项目甲", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        [None, "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 30, ""],
        [None, "", "", "", "", "P2 配置评估", "2026-07-11", "2026-07-20", "", "", "王五", "未开始", "", ""],
        [2, "新需求", "项目乙", "赵六", "中东区", "", "", "", "", "", "", "", "", ""],
        [None, "", "", "", "", "P4 工业设计", "2026-08-01", "2026-08-30", "", "", "钱七", "未开始", "", ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 2
    assert report.phases_imported == 3
    assert report.errors == []
    projs = db_session.query(Project).order_by(Project.id).all()
    assert len(projs[0].phases) == 2   # 甲：2 阶段
    assert len(projs[1].phases) == 1   # 乙：1 阶段


def test_phase_without_parent_project_is_error(client, db_session):
    """用例 3：文件首行即阶段行 → errors「缺少父项目」（评审处置 #7）。"""
    data = _make_workbook([
        [None, "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 30, ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 0
    assert report.phases_imported == 0
    assert len(report.errors) == 1
    assert "缺少父项目" in report.errors[0].message


def test_unrecognized_row_warned_not_silent(client, db_session):
    """用例 9：编号 'abc' 无阶段类型有备注 → 警告 + 跳过，不静默。"""
    data = _make_workbook([
        [1, "新需求", "正常项目", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        ["abc", "", "", "", "", "", "", "", "", "", "", "", "", "这行备注"],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 1
    assert report.phases_imported == 0
    assert report.errors == []
    assert len(report.warnings) == 1
    assert "无法识别" in report.warnings[0].message
    assert "abc" in report.warnings[0].message


def test_unmappable_phase_type_warned_and_skipped(client, db_session):
    """用例 10：'自定义阶段' 不在映射表 → 警告 + 跳过该行（§2.6 决策 2）。"""
    data = _make_workbook([
        [1, "新需求", "类型测试项目", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "自定义阶段", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 30, ""],
        ["1-2", "", "", "", "", "P1 需求评估", "2026-07-11", "2026-07-20", "", "", "王五", "未开始", "", ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 1
    assert report.phases_imported == 1          # 仅 1-2 落库
    assert any("无法识别" in w.message and "自定义阶段" in w.message
               for w in report.warnings)
    phases = db_session.query(Phase).all()
    assert [ph.phase_type for ph in phases] == ["P1"]


def test_parent_code_mismatch_warns_and_attaches_to_current(client, db_session):
    """用例 11：阶段行 9.1 但当前项目（文件内编号）为 3 → 警告 + 归属 current。"""
    data = _make_workbook([
        [3, "新需求", "当前项目", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        ["9.1", "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 30, ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 1
    assert report.phases_imported == 1
    assert any("不一致" in w.message for w in report.warnings)
    proj = db_session.query(Project).one()
    assert len(proj.phases) == 1                # 仍归属当前项目


def test_pure_digit_code_with_phase_type_is_phase_row(client, db_session):
    """用例 13：阶段行编号 '1' + 阶段类型 'P1 需求评估' → 阶段行归属 current，不误建项目。"""
    data = _make_workbook([
        [2, "新需求", "真项目", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        [1, "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 30, ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 1        # 编号 1 未被误判为项目
    assert report.phases_imported == 1
    proj = db_session.query(Project).one()
    assert len(proj.phases) == 1


def test_progress_formats_end_to_end(client, db_session):
    """端到端：进度多种写法统一 0-100 落库。"""
    data = _make_workbook([
        [1, "新需求", "进度项目", "张三", "拉美区", "", "", "", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "P1 需求评估", "", "", "", "", "李四", "进行中", "50％", ""],
        ["1-2", "", "", "", "", "P2 配置评估", "", "", "", "", "王五", "进行中", "50 %", ""],
        ["1-3", "", "", "", "", "P4 工业设计", "", "", "", "", "赵六", "进行中", 0.5, ""],
        ["1-4", "", "", "", "", "P5 结构设计", "", "", "", "", "钱七", "进行中", "50", ""],
        ["1-5", "", "", "", "", "P6 样机打样", "", "", "", "", "孙八", "未开始", 0, ""],
    ])
    report = import_excel(db_session, data)
    assert report.phases_imported == 5
    assert report.errors == []
    phases = db_session.query(Phase).order_by(Phase.sequence).all()
    assert [ph.progress for ph in phases] == [50, 50, 50, 50, 0]


# ---------- 用例 12：同构 fixture + 真实表（评审处置 #8） ----------

def test_fixture_same_structure_as_real_file(client, db_session):
    """fixture：点分编号 / 空编号 / 全角 / 短日期 同构数据全部导入，0 错误。"""
    data = _make_workbook([
        # 项目块 1：点分编号 + 全角日期
        [1, "新需求", "Fixture项目一", "张三", "拉美区", "", "2026.6.29", "2026.9.30", "", "", "", "", "", ""],
        ["1.1", "", "", "", "", "P1 需求评估", "２０２６．６．２９", "2026.8.30", "", "", "李四", "已完成", 100, ""],
        [None, "", "", "", "", "P2 配置评估", "2026年7月1日", "2026年7月15日", "", "", "王五", "进行中", "50％", ""],
        # 项目块 2：短日期 + 空编号阶段
        [2, "量产", "Fixture项目二", "赵六", "非洲区", "", "2026.3.10", "2026.6.30", "", "", "", "", "", ""],
        [None, "", "", "", "", "P1 需求评估", "6.29", "7.30", "", "", "钱七", "未开始", "", ""],
        # 项目块 3：纯数字编号 + 阶段类型（评审 #1）
        [3, "新需求", "Fixture项目三", "孙八", "东欧区", "", "", "", "", "", "", "", "", ""],
        [1, "", "", "", "", "P5 结构设计", "", "", "", "", "周九", "进行中", 0.5, ""],
    ])
    report = import_excel(db_session, data)
    assert report.projects_imported == 3
    assert report.phases_imported == 4
    assert report.errors == []
    projs = db_session.query(Project).order_by(Project.id).all()
    assert [len(p.phases) for p in projs] == [2, 1, 1]
    # 短日期 → 当年 + 警告可见（决策 A）
    assert any("缺少年份" in w.message for w in report.warnings)
    # 全角/点分/中文日期全部解析成功（无"无法解析日期"）
    assert not any("无法解析日期" in w.message for w in report.warnings)


@pytest.fixture()
def real_excel_bytes():
    if not _REAL_EXCEL.exists():
        pytest.skip(f"真实 Excel 文件不存在: {_REAL_EXCEL}")
    return _REAL_EXCEL.read_bytes()


def test_real_user_workbook_full_import(client, db_session, real_excel_bytes):
    """用例 12：真实表全量导入——核心不变量（项目数/阶段数随用户文件漂移，断言用下限）。

    背景（§一）：改造前 206 行仅 38 项目 + 4 阶段，164 行静默丢失；改造后 0 丢失。
    （2026-08-27 用户文件更新为 269 行/36 项目/164 阶段——断言项目≥30、阶段≥160，
     精确值随用户文件变动会过时。）
    剩余警告均为 '待定' 等占位文本日期（§2.4 失败 → 警告设空，预期行为），
    无任何"已跳过"行（数据全部落库）。
    """
    report = import_excel(db_session, real_excel_bytes)
    assert report.projects_imported >= 30
    assert report.phases_imported >= 160
    assert report.errors == []
    assert not any("已跳过" in w.message for w in report.warnings)
    assert db_session.query(Project).count() == report.projects_imported
    assert db_session.query(Phase).count() == report.phases_imported


# ---------- 用例 14：合并导入比较键归一化（评审处置 #2） ----------

def test_merge_import_fullwidth_name_matches(client, db_session):
    """同名项目先以全角名称入库，再用半角名称文件合并导入 → 命中合并而非新增。"""
    from app.services.excel_importer import import_merged
    # 先以全角名称入库（模拟历史数据 NFKC 差异）
    data_fullwidth = _make_workbook([
        [1, "新需求", "ＡＢＣ银行定制机", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 50, ""],
    ])
    report1 = import_merged(db_session, parse_workbook(data_fullwidth))
    assert report1.projects_created == 1

    # 再用半角名称合并导入 → 归一化后命中（R2 更新），不新增
    data_halfwidth = _make_workbook([
        [1, "新需求", "ABC银行定制机", "王五", "中东区", "", "2026-07-01", "2026-09-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "P1 需求评估", "2026-07-01", "2026-07-15", "", "", "李四", "已完成", 100, ""],
    ])
    report2 = import_merged(db_session, parse_workbook(data_halfwidth))
    assert report2.projects_updated == 1      # 命中合并
    assert report2.projects_created == 0      # 而非新增
    assert db_session.query(Project).count() == 1
    proj = db_session.query(Project).one()
    db_session.refresh(proj)
    assert proj.owner == "王五"               # 合并更新生效


# ---------- 用例 15：多人字段拆分（顿号/逗号/空格混合） ----------

def test_split_persons_dunhao_separated():
    """阶段负责人 '张三、李四'（顿号分隔）应拆为两人，而非整体当作一人。"""
    from app.services.excel_importer import split_persons
    assert split_persons("张三、李四") == ["张三", "李四"]
    assert split_persons("张三，李四 王五") == ["张三", "李四", "王五"]
    assert split_persons("张三、张三、李四") == ["张三", "李四"]  # 去重保序
    assert split_persons("/") == []
    assert split_persons("") == []
    assert split_persons(None) == []
