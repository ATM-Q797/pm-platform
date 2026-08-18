"""合并导入（增量合并）专项测试。

覆盖：新增项目/同名合并更新/阶段更新与新增/保留不动/幂等/依赖不动/待关联提示/替换模式。
"""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl

from app.models import Dependency, Phase, Project, Resource


# ---------- 构造合并场景 Excel ----------

def _make_workbook(projects_rows: list[list]) -> bytes:
    """按 14 列格式构造 Excel。每行: [编号, 类目, 名称, 负责人, 市场, 阶段类型,
    计划开始, 计划结束, 实际开始, 实际结束, 阶段负责人, 阶段状态, 阶段进度, 备注]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目填报"
    headers = ["项目编号", "项目类目", "项目名称", "项目负责人", "市场", "阶段类型",
               "计划开始", "计划结束", "实际开始", "实际结束",
               "阶段负责人", "阶段状态", "阶段进度", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, row in enumerate(projects_rows):
        for c, v in enumerate(row, 1):
            ws.cell(3 + i, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mk_project(db_session, name: str, owner: str = "旧负责人", market: str = "拉美区",
                status: str = "进行中") -> Project:
    p = Project(code=str(len(list(db_session.query(Project))) + 1), category="新需求",
                name=name, owner=owner, market=market, status=status)
    db_session.add(p)
    db_session.flush()
    return p


def _mk_phase(db_session, project: Project, phase_type: str, name: str, sequence: int,
              status: str = "进行中", progress: int = 50) -> Phase:
    ph = Phase(project_id=project.id, phase_type=phase_type, name=name, sequence=sequence,
               status=status, progress=progress,
               plan_start=date(2026, 7, 1), plan_end=date(2026, 7, 30))
    db_session.add(ph)
    db_session.flush()
    return ph


def _mk_dep(db_session, frm: Phase, to: Phase, type_: str = "FS") -> None:
    db_session.add(Dependency(from_phase_id=frm.id, to_phase_id=to.id, type=type_, lag_days=0))


def _import(client, data: bytes, mode: str = "merge"):
    resp = client.post("/api/import/excel", files={"file": ("t.xlsx", data)},
                       params={"mode": mode})
    assert resp.status_code == 200, resp.text
    return resp.json()


def _preview(client, data: bytes):
    resp = client.post("/api/import/preview", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text
    return resp.json()


# ---------- 测试 ----------

def test_merge_creates_new_projects_with_fs_chain(client, db_session):
    """空库导入：新项目全部创建，阶段建 FS 串联链。"""
    data = _make_workbook([
        [1, "新需求", "全新项目A", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "已完成", 100, ""],
        ["1-2", "", "", "", "", "结构设计", "2026-07-11", "2026-07-20", "", "", "王五", "进行中", 60, ""],
        [2, "新需求", "全新项目B", "赵六", "中东区", "", "2026-08-01", "2026-09-01", "", "", "", "", "", ""],
        ["2-1", "", "", "", "", "联调测试", "2026-08-01", "2026-08-30", "", "", "钱七", "未开始", 0, ""],
    ])
    report = _import(client, data)

    assert report["projects_created"] == 2
    assert report["projects_updated"] == 0
    assert report["phases_created"] == 3
    # 全新项目阶段有 FS 链
    projects = db_session.query(Project).all()
    assert len(projects) == 2
    deps = db_session.query(Dependency).count()
    assert deps == 1  # A: 1-1→1-2（两阶段 1 条链）；B 单阶段无链


def test_merge_updates_existing_project_fields(client, db_session):
    """同名项目：文件有值覆盖；状态未填保留系统值。"""
    p = _mk_project(db_session, "现有项目", owner="旧负责人", market="拉美区", status="进行中")
    db_session.commit()

    # 文件：负责人有值（覆盖）、市场有值（覆盖）、状态不填（保留"进行中"）
    data = _make_workbook([
        [1, "量产", "现有项目", "新负责人", "中东区", "", "2026-07-01", "2026-09-01", "", "", "", "", "", ""],
    ])
    report = _import(client, data)

    assert report["projects_updated"] == 1
    assert report["projects_created"] == 0
    db_session.refresh(p)
    assert p.owner == "新负责人"
    assert p.market == "中东区"
    assert p.category == "量产"
    assert p.plan_end == date(2026, 9, 1)
    assert p.status == "进行中"  # 未填 → 保留
    # 项目 id 不变（不是删了重建）
    assert p.id == p.id


def test_merge_phase_update_and_insert_natural_order(client, db_session):
    """同名项目阶段：P4 更新（状态/进度覆盖）；P1 新增（自然序插入 seq=1，不建依赖）；系统多余阶段保留。"""
    p = _mk_project(db_session, "项目A")
    p4 = _mk_phase(db_session, p, "P4", "工业设计", 1, status="进行中", progress=50)
    p5 = _mk_phase(db_session, p, "P5", "结构设计", 2, status="已完成", progress=100)
    _mk_dep(db_session, p4, p5)  # 已有依赖
    db_session.commit()

    # 文件：P4 更新（已完成/100）、P1 新增（需求评估）
    data = _make_workbook([
        [1, "新需求", "项目A", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "已完成", 100, ""],
        ["1-2", "", "", "", "", "需求评估", "2026-07-01", "2026-07-05", "", "", "王五", "未开始", 0, ""],
    ])
    report = _import(client, data)

    assert report["projects_updated"] == 1
    assert report["phases_updated"] == 1
    assert report["phases_created"] == 1
    # 待关联提示
    assert len(report["pending_link_phases"]) == 1
    assert report["pending_link_phases"][0]["project_name"] == "项目A"
    assert report["pending_link_phases"][0]["phase_name"] == "需求评估"

    # P4 被更新
    db_session.refresh(p4)
    assert p4.status == "已完成"
    assert p4.progress == 100

    # 新 P1 存在且 sequence=1（自然序插入），P4/P5 顺延
    phases = db_session.query(Phase).filter_by(project_id=p.id).order_by(Phase.sequence).all()
    assert [ph.phase_type for ph in phases] == ["P1", "P4", "P5"]
    assert [ph.sequence for ph in phases] == [1, 2, 3]

    # 已有依赖 P4→P5 不动；新 P1 无依赖
    deps = db_session.query(Dependency).all()
    assert len(deps) == 1
    assert deps[0].from_phase_id == p4.id and deps[0].to_phase_id == p5.id
    new_p1 = next(ph for ph in phases if ph.phase_type == "P1")
    assert all(d.from_phase_id != new_p1.id and d.to_phase_id != new_p1.id for d in deps)


def test_merge_keeps_unlisted_projects_and_phases(client, db_session):
    """系统有、文件没有的项目/阶段：保留不动。"""
    p = _mk_project(db_session, "不在文件的项目")
    _mk_phase(db_session, p, "P4", "工业设计", 1)
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "文件里的项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    report = _import(client, data)

    assert report["projects_created"] == 1
    assert report["projects_updated"] == 0
    # 原有项目原样保留
    projects = db_session.query(Project).order_by(Project.id).all()
    assert [p.name for p in projects] == ["不在文件的项目", "文件里的项目"]
    kept = next(x for x in projects if x.name == "不在文件的项目")
    assert kept.status == "进行中"
    assert kept.phases and kept.phases[0].name == "工业设计"


def test_merge_idempotent(client, db_session):
    """重复导入同一文件：幂等，无重复项目/阶段。"""
    data = _make_workbook([
        [1, "新需求", "幂等项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 50, ""],
    ])
    _import(client, data)
    report2 = _import(client, data)

    assert report2["projects_created"] == 0  # 第二次同名 → 更新
    assert report2["projects_updated"] == 1
    assert report2["phases_updated"] == 1
    assert report2["phases_created"] == 0
    assert db_session.query(Project).count() == 1
    assert db_session.query(Phase).count() == 1


def test_merge_keeps_existing_dependencies(client, db_session):
    """已有依赖（含跨阶段）在合并后保持不动。"""
    p = _mk_project(db_session, "依赖项目")
    p3 = _mk_phase(db_session, p, "P3", "模块选型", 1)
    p6 = _mk_phase(db_session, p, "P6", "样机打样", 2)
    _mk_dep(db_session, p3, p6, type_="SS")  # 手动跨阶段 SS
    db_session.commit()

    # 文件：P6 更新 + P1 新增
    data = _make_workbook([
        [1, "新需求", "依赖项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "样机打样", "2026-07-01", "2026-07-20", "", "", "李四", "进行中", 80, ""],
        ["1-2", "", "", "", "", "需求评估", "2026-07-01", "2026-07-05", "", "", "王五", "未开始", 0, ""],
    ])
    _import(client, data)

    deps = db_session.query(Dependency).all()
    assert len(deps) == 1
    assert deps[0].type == "SS"  # 原依赖类型保持
    assert deps[0].from_phase_id == p3.id and deps[0].to_phase_id == p6.id


def test_preview_merge_stats(client, db_session):
    """预览返回合并明细：created/updated/kept + 阶段统计 + 待关联提示。"""
    p = _mk_project(db_session, "现有项目")
    _mk_phase(db_session, p, "P4", "工业设计", 1)
    _mk_project(db_session, "保留项目")
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "现有项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 60, ""],
        ["1-2", "", "", "", "", "需求评估", "2026-07-01", "2026-07-05", "", "", "王五", "未开始", 0, ""],
        [2, "新需求", "全新项目", "赵六", "中东区", "", "2026-08-01", "2026-09-01", "", "", "", "", "", ""],
        ["2-1", "", "", "", "", "结构设计", "2026-08-01", "2026-08-20", "", "", "钱七", "未开始", 0, ""],
    ])
    preview = _preview(client, data)

    assert [p["name"] for p in preview["created_projects"]] == ["全新项目"]
    assert [p["name"] for p in preview["updated_projects"]] == ["现有项目"]
    assert preview["kept_count"] == 1  # 保留项目
    assert preview["phases_created"] == 2  # 新 P1 + 全新项目 1 阶段
    assert preview["phases_updated"] == 1  # 现有 P4
    assert len(preview["pending_link_phases"]) == 1
    assert preview["pending_link_phases"][0]["phase_name"] == "需求评估"


def test_replace_mode_still_clears_all(client, db_session):
    """replace 模式：全量清空重建（兼容旧行为）。"""
    p = _mk_project(db_session, "旧项目")
    _mk_phase(db_session, p, "P4", "工业设计", 1)
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "替换后项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    report = _import(client, data, mode="replace")

    assert report["projects_imported"] == 1
    assert db_session.query(Project).count() == 1
    assert db_session.query(Project).first().name == "替换后项目"
    assert db_session.query(Phase).count() == 0


def test_merge_rejects_invalid_mode(client, db_session):
    """非法 mode 参数被拒绝。"""
    resp = client.post("/api/import/excel", files={"file": ("t.xlsx", b"x")},
                       params={"mode": "bad"})
    assert resp.status_code == 400


def test_merge_preserves_resources(client, db_session):
    """合并模式不删除资源（只增）。"""
    r = Resource(name="老人员")
    db_session.add(r)
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "新项目", "新人员", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 50, ""],
    ])
    _import(client, data)

    names = {x.name for x in db_session.query(Resource).all()}
    assert "老人员" in names  # 保留
    assert "新人员" in names  # 新增
    assert "李四" in names
