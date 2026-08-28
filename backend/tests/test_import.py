"""Excel 导入器测试。

分两类：
1. 单元测试：清洗工具函数（parse_cell_date / split_persons / classify_row / map_phase_type）
2. 端到端测试：用真实 Excel 文件做导入（文件缺失则 skip）

真实文件路径：~/Desktop/整机项目进度及计划情况统计-0720.xlsx
预期：国内 8 + 海外 10 = 18 项目，57 阶段。
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest

from app.models import Dependency, Phase, Project, Resource
from app.schemas.import_report import ImportReport
from app.services.excel_importer import (
    PHASE_NAME_TO_TYPE,
    classify_row,
    import_excel,
    map_phase_type,
    parse_cell_date,
    split_persons,
)

_REAL_EXCEL = Path.home() / "Desktop" / "整机项目进度及计划情况统计-0720.xlsx"


# ---------- 单元测试：清洗函数 ----------

def _empty_report() -> ImportReport:
    return ImportReport()


def test_classify_row():
    assert classify_row("1", "项目A") == "project"
    assert classify_row("1-1", "工业设计") == "phase"
    assert classify_row("2-3", "结构设计") == "phase"
    assert classify_row(None, None) == "skip"           # 空行
    assert classify_row("【状态图例】", None) == "skip"  # 备注行
    assert classify_row("未发货样机", "长文本...") == "skip"


def test_split_persons_multi_space():
    """多空格分隔的多人字段。"""
    assert split_persons("苏树冲             陆永杰") == ["苏树冲", "陆永杰"]


def test_split_persons_variants():
    """逗号（中英文）、换行、"/" 占位符。"""
    assert split_persons("曹俊杰,宋海波") == ["曹俊杰", "宋海波"]
    assert split_persons("曹俊杰，宋海波") == ["曹俊杰", "宋海波"]
    assert split_persons("曹俊杰\n宋海波") == ["曹俊杰", "宋海波"]
    assert split_persons("/") == []
    assert split_persons(None) == []
    assert split_persons("") == []
    # 去重
    assert split_persons("张三 张三") == ["张三"]


def test_parse_cell_date_datetime():
    """datetime 对象 → date。"""
    report = _empty_report()
    d = parse_cell_date(datetime(2026, 7, 1, 12, 0), 1, "国内", "plan_start", report)
    assert d == date(2026, 7, 1)
    assert report.warnings == []  # 正常解析无 warning


def test_parse_cell_date_anomaly_text():
    """文本异常（'2026/-/--'）→ None + warning。"""
    report = _empty_report()
    d = parse_cell_date("2026/-/--", 31, "国内", "plan_end", report)
    assert d is None
    assert len(report.warnings) == 1
    assert "无法解析日期" in report.warnings[0].message


def test_parse_cell_date_none():
    assert parse_cell_date(None, 1, "x", "f", _empty_report()) is None
    assert parse_cell_date("", 1, "x", "f", _empty_report()) is None


def test_map_phase_type_known():
    """文档 §5.3 映射表内的阶段名（PHASE_TYPES_V2 §二 重排后编号）。"""
    report = _empty_report()
    assert map_phase_type("工业设计", 1, "国内", report) == "P4"
    assert map_phase_type("结构设计", 1, "国内", report) == "P5"
    assert map_phase_type("线缆设计", 1, "国内", report) == "P6"   # 新环节
    assert map_phase_type("样机打样", 1, "国内", report) == "P71"  # 原 P6
    assert map_phase_type("线缆打样", 1, "国内", report) == "P72"  # 新增
    assert map_phase_type("联调测试", 1, "国内", report) == "P8"   # 原 P7
    assert map_phase_type("POC及投标", 1, "国内", report) == "P9"  # 原 P8 交付族
    assert map_phase_type("交付", 1, "国内", report) == "P9"
    assert map_phase_type("需求分析", 1, "国内", report) == "P1"
    assert report.errors == []


def test_map_phase_type_unknown_records_error():
    """不在映射表的阶段名 → None + error。"""
    report = _empty_report()
    # 用真正不存在的阶段名（扩充映射表后已覆盖实际 Excel 的所有变体）
    assert map_phase_type("完全不存在的阶段XYZ", 1, "海外", report) is None
    assert map_phase_type("未知阶段", 1, "海外", report) is None
    assert len(report.errors) == 2


def test_phase_type_rank_renumbering():
    """PHASE_TYPES_V2 §八 用例 2/3b：排序 P6 < P71 < P72 < P8 < P9；
    旧值归一——旧 P7（联调测试）排 P72 之后（=新 P8 位）、旧 P8（交付）= 新 P9 位；未知排最后。"""
    from app.services.excel_importer import _phase_type_rank

    order = ["P1", "P2", "P3", "P4", "P5", "P6", "P71", "P72", "P8", "P9"]
    ranks = [_phase_type_rank(pt) for pt in order]
    assert ranks == sorted(ranks), "新编号排序错乱"
    # 旧值归一（决策 ③ 历史数据不迁移）
    assert _phase_type_rank("P7") == (8, 0)  # 旧联调测试 → 新 P8 位（排 P72 之后）
    # P8 → (9, 0)：旧交付归一到 P9 位；新 P8=联调测试 phase_type 相同、机制上无法区分，
    # 排序同样落 (9, 0) 与 P9 同位——与 §四排除的「机制双兼容」同源，设计 §三 明文接受
    assert _phase_type_rank("P8") == _phase_type_rank("P9") == (9, 0)
    assert _phase_type_rank("P6") < _phase_type_rank("P71")   # 旧样机打样在 P71 之前
    assert _phase_type_rank("线缆设计") == _phase_type_rank("ZZZ") == (99, 99)  # 未知排最后


def test_phase_name_mapping_table_completeness():
    """验证映射表覆盖了文档 §5.3 原表 + 新环节 + 实际 Excel 的 5 个变体（PHASE_TYPES_V2 §二）。"""
    # §5.3 原表 18 个 + 线缆设计/线缆打样（新增环节）
    expected_keys = {
        "工业设计", "结构设计", "整机设计", "样机打样", "线缆设计", "线缆打样",
        "联调测试", "测试", "POC及投标",
        "需求分析", "需求评估", "配置评估", "模块选型",
        "直接投料", "图纸归档", "归档", "BOM制作与激活",
        "首批生产保障", "投料", "发货",
    }
    # 实际 Excel 的 6 个变体（扩充部分）
    expanded_keys = {
        "测试与发货", "样机打样（1台）", "归档（归档后再投料）",
        "直接投料，BOM制作与激活", "直接投料，激活时间", "交付",
    }
    all_keys = expected_keys | expanded_keys
    assert all_keys <= set(PHASE_NAME_TO_TYPE.keys()), (
        f"映射表缺少: {all_keys - set(PHASE_NAME_TO_TYPE.keys())}"
    )


# ---------- 端到端测试：真实 Excel 导入 ----------

@pytest.fixture()
def real_excel_bytes():
    if not _REAL_EXCEL.exists():
        pytest.skip(f"真实 Excel 文件不存在: {_REAL_EXCEL}")
    return _REAL_EXCEL.read_bytes()


def test_import_real_excel(client, db_session, real_excel_bytes):
    """导入真实 Excel：应得到 18 项目（国内 8 + 海外 10）。"""
    resp = client.post(
        "/api/import/excel",
        files={"file": ("整机项目进度及计划情况统计-0720.xlsx", real_excel_bytes)},
    )
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["projects_imported"] == 18
    assert report["phases_imported"] == 57

    # 校验市场分布
    projects = client.get("/api/projects").json()
    domestic = [p for p in projects if p["market"] == "国内"]
    overseas = [p for p in projects if p["market"] == "海外"]
    assert len(domestic) == 8
    assert len(overseas) == 10
    # 项目类目默认为"新需求"
    assert all(p["category"] == "新需求" for p in projects)


def test_import_idempotent(client, db_session, real_excel_bytes):
    """重复导入幂等：第二次导入后仍是 18 项目。"""
    for _ in range(2):
        resp = client.post(
            "/api/import/excel",
            files={"file": ("test.xlsx", real_excel_bytes)},
        )
        assert resp.status_code == 200
    assert resp.json()["projects_imported"] == 18
    assert len(client.get("/api/projects").json()) == 18


def test_import_creates_resources(client, db_session, real_excel_bytes):
    """多人字段被正确拆分为独立人员。"""
    client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", real_excel_bytes)},
    )
    resources = client.get("/api/resources").json()
    # 至少有若干人员
    assert len(resources) >= 10
    names = {r["name"] for r in resources}
    # 验证某些已知人员存在（基于数据探查）
    assert "曹俊杰" in names or "何昊" in names


def test_import_date_anomalies_recorded_as_warnings(client, db_session, real_excel_bytes):
    """'2026/-/--' 类日期被记为 warning，对应 plan_end/plan_start 为 null。"""
    resp = client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", real_excel_bytes)},
    )
    report = resp.json()
    # 数据探查发现多处日期异常，导入时被正确捕获为 warning
    assert len(report["warnings"]) >= 3
    # 至少有一个阶段/项目的日期被设为 null（埃塞 AWASH 项目 plan_start='2026/-/--'）
    has_null_date = False
    for p in client.get("/api/projects").json():
        if p["plan_start"] is None or p["plan_end"] is None:
            has_null_date = True
            break
    assert has_null_date


def test_import_unmappable_phase_names_recorded_as_errors(client, db_session, real_excel_bytes):
    """扩充映射表后，实际 Excel 的所有阶段名都能映射，无 error。"""
    resp = client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", real_excel_bytes)},
    )
    report = resp.json()
    # 映射表已覆盖所有实际阶段名（含 测试与发货、样机打样（1台）等变体）
    assert len(report["errors"]) == 0
    # 阶段总数仍是 57
    assert report["phases_imported"] == 57


def test_import_builds_dependencies(client, db_session, real_excel_bytes):
    """导入后每个项目的阶段按 sequence 建了 FS 依赖。"""
    client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", real_excel_bytes)},
    )
    # 取第一个项目的甘特图，验证有 link
    pid = client.get("/api/projects").json()[0]["id"]
    gantt = client.get(f"/api/projects/{pid}/gantt").json()
    assert len(gantt["links"]) >= 1
    assert all(l["type"] == "0" for l in gantt["links"])  # FS


def test_import_report_endpoint(client, db_session, real_excel_bytes):
    """GET /api/import/report 能取到最近一次导入报告。"""
    client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", real_excel_bytes)},
    )
    resp = client.get("/api/import/report")
    assert resp.status_code == 200
    assert resp.json()["projects_imported"] == 18


def test_import_report_404_when_none(client, db_session):
    """无导入记录时 GET /api/import/report 返回 404。"""
    resp = client.get("/api/import/report")
    assert resp.status_code == 404


def test_import_rejects_non_excel(client, db_session):
    """非 Excel 文件被拒绝。"""
    resp = client.post(
        "/api/import/excel",
        files={"file": ("test.txt", b"hello")},
    )
    assert resp.status_code == 400



# ---------- 导入前差异报告（预览） ----------

def _make_preview_workbook() -> bytes:
    """构造 14 列模板格式 Excel（2 项目 / 3 阶段），用于预览测试。"""
    import openpyxl
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目填报"
    headers = ["项目编号", "项目类目", "项目名称", "项目负责人", "市场", "阶段类型",
               "计划开始", "计划结束", "实际开始", "实际结束",
               "阶段负责人", "阶段状态", "阶段进度", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    # 项目 1（与库中已有"现有项目A"同名）
    ws.cell(3, 1, "1")
    ws.cell(3, 3, "现有项目A")
    ws.cell(3, 5, "拉美区")
    ws.cell(3, 7, datetime(2026, 7, 1))
    ws.cell(3, 8, datetime(2026, 8, 1))
    ws.cell(4, 1, "1-1")
    ws.cell(4, 6, "工业设计")
    ws.cell(4, 7, datetime(2026, 7, 1))
    ws.cell(4, 8, datetime(2026, 7, 10))
    ws.cell(4, 11, "李四")
    ws.cell(4, 12, "已完成")
    ws.cell(4, 13, 100)
    # 项目 2（全新项目）
    ws.cell(5, 1, "2")
    ws.cell(5, 3, "全新项目B")
    ws.cell(5, 5, "中东区")
    ws.cell(6, 1, "2-1")
    ws.cell(6, 6, "结构设计")
    ws.cell(6, 7, datetime(2026, 8, 1))
    ws.cell(6, 8, datetime(2026, 8, 30))
    ws.cell(6, 11, "王五")
    ws.cell(6, 12, "未开始")

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_preview_no_side_effect(client, db_session):
    """预览不落库：调用后现有数据保持不变。"""
    # 先建一个项目
    r = client.post("/api/projects", json={
        "category": "新需求", "name": "现有项目A", "owner": "张三", "market": "拉美区",
    })
    assert r.status_code == 201

    data = _make_preview_workbook()
    resp = client.post("/api/import/preview", files={"file": ("test.xlsx", data)})
    assert resp.status_code == 200, resp.text
    preview = resp.json()

    # 现有数据统计
    assert preview["existing"]["projects"] == 1
    assert preview["existing"]["phases"] == 0

    # 文件数据统计
    assert preview["incoming"]["projects"] == 2
    assert preview["incoming"]["phases"] == 2

    # 同名对比：1 个同名匹配 / 1 个新增 / 0 个缺失
    assert preview["match"]["matched"] == 1
    assert preview["match"]["new"] == 1
    assert preview["match"]["missing"] == 0

    # 项目概览
    assert len(preview["projects_preview"]) == 2
    assert preview["projects_preview"][0]["phases"] == 1

    # 无错误无警告
    assert preview["errors"] == []
    assert preview["warnings"] == []

    # 关键：预览后库中数据未变（项目仍在）
    projects = client.get("/api/projects").json()
    assert len(projects) == 1
    assert projects[0]["name"] == "现有项目A"


def test_preview_missing_projects_detected(client, db_session):
    """同名对比：现有项目不在文件中时 missing 正确计数（防传错文件）。"""
    client.post("/api/projects", json={
        "category": "新需求", "name": "库中项目X", "owner": "张三", "market": "拉美区",
    })
    client.post("/api/projects", json={
        "category": "新需求", "name": "库中项目Y", "owner": "李四", "market": "中东区",
    })

    data = _make_preview_workbook()
    resp = client.post("/api/import/preview", files={"file": ("test.xlsx", data)})
    preview = resp.json()

    # 文件只有 1 个与库同名（现有项目A 在库中不存在了——因为库中是 X/Y）
    assert preview["match"]["matched"] == 0
    assert preview["match"]["new"] == 2
    assert preview["match"]["missing"] == 2  # X、Y 都不在文件中


def test_preview_invalid_file(client, db_session):
    """非法文件：返回错误且不落库。"""
    resp = client.post("/api/import/preview", files={"file": ("bad.xlsx", b"not an excel")})
    assert resp.status_code == 200
    preview = resp.json()
    assert len(preview["errors"]) == 1
    assert preview["incoming"]["projects"] == 0

    # 库中没有项目（无副作用）
    assert client.get("/api/projects").json() == []


def test_preview_rejects_wrong_extension(client, db_session):
    """非 Excel 扩展名被拒绝。"""
    resp = client.post("/api/import/preview", files={"file": ("a.txt", b"hello")})
    assert resp.status_code == 400


def test_preview_matches_import_result(client, db_session):
    """一致性：预览统计与确认导入后的实际结果一致。"""
    data = _make_preview_workbook()
    preview_resp = client.post("/api/import/preview", files={"file": ("test.xlsx", data)})
    preview = preview_resp.json()

    import_resp = client.post("/api/import/excel", files={"file": ("test.xlsx", data)})
    report = import_resp.json()

    assert preview["incoming"]["projects"] == report["projects_imported"]
    assert preview["incoming"]["phases"] == report["phases_imported"]

    # 导入后库中数据与预览一致
    projects = client.get("/api/projects").json()
    assert len(projects) == preview["incoming"]["projects"]


def test_import_with_user_linked_resource(client, db_session):
    """用户账户关联了 resource 时导入仍成功：先解除引用再删资源（PG 外键约束回归）。"""
    from app.models import Resource, User

    # 建 resource + 用户关联它
    res = Resource(name="关联人员")
    db_session.add(res)
    db_session.flush()
    from app.core.security import hash_password
    db_session.add(User(
        username="linkeduser",
        name="关联用户",
        role="manager",
        password_hash=hash_password("testpass"),
        resource_id=res.id,
    ))
    db_session.commit()

    data = _make_preview_workbook()
    resp = client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", data)},
        params={"mode": "replace"},
    )
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["projects_imported"] == 2

    # 用户仍在，resource_id 已被解除（资源被清空重建）
    users = db_session.query(User).all()
    assert any(u.username == "linkeduser" for u in users)
    assert all(u.resource_id is None for u in users)


# ---------- 直接调用 import_excel 的单元测试（构造内存 Excel） ----------

def test_import_excel_direct_with_constructed_workbook(client, db_session):
    """构造一个最小 14 列模板格式 Excel，直接测导入逻辑。"""
    import openpyxl
    from app.services.excel_importer import import_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目填报"
    # 表头（14 列，与 docs/项目填报模板.xlsx 一致）
    headers = ["项目编号", "项目类目", "项目名称", "项目负责人", "市场", "阶段类型",
               "计划开始", "计划结束", "实际开始", "实际结束",
               "阶段负责人", "阶段状态", "阶段进度", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    # 项目 1
    ws.cell(3, 1, "1")
    ws.cell(3, 2, "新需求")
    ws.cell(3, 3, "测试项目")
    ws.cell(3, 4, "张三")
    ws.cell(3, 5, "拉美区")
    ws.cell(3, 7, datetime(2026, 7, 1))
    ws.cell(3, 8, datetime(2026, 8, 1))
    ws.cell(3, 14, "项目备注")
    # 阶段 1-1
    ws.cell(4, 1, "1-1")
    ws.cell(4, 6, "P4 工业设计")
    ws.cell(4, 7, datetime(2026, 7, 1))
    ws.cell(4, 8, datetime(2026, 7, 10))
    ws.cell(4, 9, datetime(2026, 7, 1))
    ws.cell(4, 11, "李四 王五")
    ws.cell(4, 12, "已完成")
    ws.cell(4, 13, 100)
    # 阶段 1-2
    ws.cell(5, 1, "1-2")
    ws.cell(5, 6, "P5 结构设计")
    ws.cell(5, 7, datetime(2026, 7, 11))
    ws.cell(5, 8, datetime(2026, 7, 30))
    ws.cell(5, 11, "李四")
    ws.cell(5, 12, "进行中")
    ws.cell(5, 13, 60)
    # 阶段 1-3（纯中文阶段类型列，走映射表兜底）
    ws.cell(6, 1, "1-3")
    ws.cell(6, 6, "交付")
    ws.cell(6, 7, datetime(2026, 8, 1))
    ws.cell(6, 8, datetime(2026, 8, 30))
    ws.cell(6, 11, "王五")
    ws.cell(6, 12, "未开始")

    import io
    buf = io.BytesIO()
    wb.save(buf)

    from app.database import SessionLocal
    report = import_excel(db_session, buf.getvalue())
    assert report.projects_imported == 1
    assert report.phases_imported == 3
    assert len(report.errors) == 0

    # 项目字段验证（市场/类目/备注来自模板列）
    projects = list(db_session.query(Project))
    assert projects[0].market == "拉美区"
    assert projects[0].category == "新需求"
    assert projects[0].remark == "项目备注"

    # 验证人员拆分：李四、王五、张三
    resources = list(db_session.query(Resource))
    names = {r.name for r in resources}
    assert {"张三", "李四", "王五"} <= names

    # 验证依赖：1-1 → 1-2 → 1-3 两条 FS
    deps = list(db_session.query(Dependency))
    assert len(deps) == 2
    assert all(d.type == "FS" for d in deps)

    # 验证 phase_type 解析（"P4 工业设计" → P4 + name=工业设计）与进度/实际日期
    phases = list(db_session.query(Phase).order_by(Phase.sequence))
    assert phases[0].phase_type == "P4"
    assert phases[0].name == "工业设计"
    assert phases[0].progress == 100
    assert phases[0].actual_start == date(2026, 7, 1)
    assert phases[1].phase_type == "P5"
    assert phases[1].name == "结构设计"
    assert phases[1].progress == 60
    # 阶段 1-3：纯中文阶段类型列 → 映射表兜底（PHASE_TYPES_V2：交付 → P9）
    assert phases[2].phase_type == "P9"
    assert phases[2].name == "交付"


def test_import_old_format_compat(client, db_session):
    """旧 8 列格式（含交接人列）仍可导入：交接人列被忽略，市场按 sheet 名推断。"""
    import openpyxl
    from app.services.excel_importer import import_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目情况统计-海外"
    headers = ["项目编号", "项目类目", "项目名称", "负责人", "计划开始", "计划结束", "状态", "交接人"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    # 项目 1
    ws.cell(3, 1, "1")
    ws.cell(3, 3, "旧格式项目")
    ws.cell(3, 4, "张三")
    ws.cell(3, 7, "进行中")
    # 阶段 1-1
    ws.cell(4, 1, "1-1")
    ws.cell(4, 3, "工业设计")
    ws.cell(4, 4, "李四")
    ws.cell(4, 7, "已完成")
    ws.cell(4, 8, "王五")  # 交接人列：应被忽略

    import io
    buf = io.BytesIO()
    wb.save(buf)

    report = import_excel(db_session, buf.getvalue())
    assert report.projects_imported == 1
    assert report.phases_imported == 1

    projects = list(db_session.query(Project))
    assert projects[0].market == "海外"  # 旧格式按 sheet 名推断
    assert projects[0].status == "进行中"

    phases = list(db_session.query(Phase))
    assert phases[0].phase_type == "P4"  # 名称映射兜底
    assert phases[0].status == "已完成"
