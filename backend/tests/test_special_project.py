"""专项项目页面测试（SPECIAL_PROJECT §七）。

覆盖：三处资源负载排除（热力/甘特/冲突）、自定义阶段类型、专项列表权限与路由顺序、
gantt remark、预警口径数据、开关切换与列表过滤、普通列表/Dashboard 隔离、
详情/阶段/项目级接口权限、列表聚合排除、角标边界（后端可测部分）。
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import openpyxl

from app.core.security import hash_password
from app.models import Phase, Project, Resource, User

# ---- 公共构造 ----


def _mk_project(db_session, name: str, status: str = "进行中", is_special: bool = False,
                plan_end: date | None = None) -> Project:
    p = Project(code=f"{name}-{len(list(db_session.query(Project))) + 1}", category="新需求",
                name=name, owner="张三", market="拉美区", status=status,
                is_special=is_special, plan_end=plan_end)
    db_session.add(p)
    db_session.flush()
    return p


def _mk_phase(db_session, project: Project, name: str, start: date | None, end: date | None,
              status: str = "进行中", phase_type: str = "P4", remark: str | None = None) -> Phase:
    ph = Phase(project_id=project.id, phase_type=phase_type, name=name, sequence=1,
               plan_start=start, plan_end=end, status=status, progress=50, remark=remark)
    db_session.add(ph)
    db_session.flush()
    return ph


def _mk_resource(db_session, name: str) -> Resource:
    r = Resource(name=name)
    db_session.add(r)
    db_session.flush()
    return r


def _mk_user(db_session, username: str, role: str, name: str = "测试用户") -> User:
    u = User(username=username, name=name, role=role,
             password_hash=hash_password("testpass"))
    db_session.add(u)
    return u


def _login(client, username: str):
    resp = client.post("/api/auth/login", json={"username": username, "password": "testpass"})
    assert resp.status_code == 200, resp.text


def _today() -> date:
    return date.today()


def _week_monday(d: date) -> date:
    """d 所在周的周一（与后端热力桶对齐规则一致）。"""
    return d - timedelta(days=d.weekday())


# ---- §七 用例 ----


def test_1_special_phase_excluded_from_all_resource_views(client, db_session):
    """用例 1：专项项目阶段 → 热力图不占格、资源甘特不显示、冲突不报（三处排除）。"""
    t = _today()
    r = _mk_resource(db_session, "专项排除人")
    # 4 个普通项目阶段同窗深度重叠（并行 4 > 3 → 若无排除报 C(5,2)=10 对）
    for i in range(4):
        p = _mk_project(db_session, f"普通项目{i + 1}")
        _mk_phase(db_session, p, f"普通阶段{i + 1}", t, t + timedelta(days=30)).assignees = [r]
    # 专项项目阶段同一窗口
    sp = _mk_project(db_session, "专项项目", is_special=True)
    _mk_phase(db_session, sp, "专项阶段", t, t + timedelta(days=30)).assignees = [r]
    db_session.commit()

    # 冲突：仅 4 个普通阶段参与 → C(4,2)=6 对，专项阶段不在任何对里
    conflicts = client.get("/api/resources/conflicts").json()
    assert len(conflicts) == 1
    pairs = conflicts[0]["conflicts"]
    assert len(pairs) == 6
    involved = {c["phase_a_id"] for c in pairs} | {c["phase_b_id"] for c in pairs}
    assert all(ph.id not in involved for ph in db_session.query(Phase).filter(Phase.name == "专项阶段"))

    # 热力图：当前周桶格值 = 4（不含专项），peak=4（30 天阶段跨 ~5 周，sum 断言按桶）
    hm = client.get("/api/resources/heatmap").json()
    row = next(p for p in hm["people"] if p["resource_id"] == r.id)
    cur_col = hm["columns"].index(_week_monday(t).isoformat())
    assert row["cells"][cur_col] == 4 and row["peak_parallel"] == 4
    phase_ids = {e["phase_id"] for cp in row["cell_phases"] if cp for e in cp}
    assert all(ph.id not in phase_ids for ph in db_session.query(Phase).filter(Phase.name == "专项阶段"))

    # 资源甘特：/all/workload 与 /{id}/workload 均不含专项阶段
    all_wl = client.get("/api/resources/all/workload").json()
    me = next(w for w in all_wl if w["resource"]["id"] == r.id)
    assert len(me["workloads"]) == 4
    assert all("专项" not in w["project_name"] for w in me["workloads"])


def test_2_custom_phase_type_saved_and_gantt_ok(client, db_session):
    """用例 2：专项项目阶段类型自由文本（如"电磁兼容测试"）→ 保存成功、甘特正常。"""
    resp = client.post("/api/projects", json={
        "name": "专项自由类型", "category": "定制", "owner": "张三", "market": "拉美区",
        "is_special": True,
    })
    assert resp.status_code == 201, resp.text
    pid = resp.json()["id"]

    t = _today()
    resp = client.post(f"/api/projects/{pid}/phases", json={
        "phase_type": "电磁兼容测试", "name": "电磁兼容测试", "sequence": 1,
        "plan_start": t.isoformat(), "plan_end": (t + timedelta(days=10)).isoformat(),
        "status": "进行中", "progress": 30,
    })
    assert resp.status_code == 201, resp.text
    assert resp.json()["phase_type"] == "电磁兼容测试"

    gantt = client.get(f"/api/projects/{pid}/gantt")
    assert gantt.status_code == 200, gantt.text
    task_texts = [tk["text"] for tk in gantt.json()["data"]]
    assert "电磁兼容测试" in task_texts


def test_4_special_list_permissions_and_route_order(client, db_session):
    """用例 4：专项列表 admin/manager 200；engineer/viewer 403；/special 不被 /{project_id} 吞掉。"""
    _mk_project(db_session, "专项路由项目", is_special=True)
    db_session.commit()

    # admin（默认登录）200 且返回专项
    resp = client.get("/api/projects/special")
    assert resp.status_code == 200, resp.text
    assert [p["name"] for p in resp.json()] == ["专项路由项目"]

    # manager 200
    _mk_user(db_session, "mgr_special", "manager", "经理甲")
    db_session.commit()
    _login(client, "mgr_special")
    resp = client.get("/api/projects/special")
    assert resp.status_code == 200, resp.text

    # engineer 403（路由未被 /{project_id} 吞掉：若被吞会 422，而非 403）
    _mk_user(db_session, "eng_special", "engineer", "工程师甲")
    db_session.commit()
    _login(client, "eng_special")
    resp = client.get("/api/projects/special")
    assert resp.status_code == 403, resp.text

    # viewer 403
    _mk_user(db_session, "view_special", "viewer", "观察员甲")
    db_session.commit()
    _login(client, "view_special")
    resp = client.get("/api/projects/special")
    assert resp.status_code == 403, resp.text


def test_5_gantt_remark_carried(client, db_session):
    """用例 5：gantt 响应 remark：阶段有备注 → task.remark 携带（项目甘特 + /all/workload）。"""
    t = _today()
    r = _mk_resource(db_session, "备注人")
    p = _mk_project(db_session, "备注项目")
    ph = _mk_phase(db_session, p, "备注阶段", t, t + timedelta(days=7), remark="注意：样机需提前")
    ph.assignees = [r]
    db_session.commit()

    gantt = client.get(f"/api/projects/{p.id}/gantt")
    assert gantt.status_code == 200
    task = next(tk for tk in gantt.json()["data"] if tk["id"] == ph.id)
    assert task["remark"] == "注意：样机需提前"

    all_wl = client.get("/api/resources/all/workload").json()
    me = next(w for w in all_wl if w["resource"]["id"] == r.id)
    item = next(w for w in me["workloads"] if w["phase_id"] == ph.id)
    assert item["remark"] == "注意：样机需提前"


def test_6_badge_input_data_available(client, db_session):
    """用例 6：预警口径相关字段（延期/即将到期/无阶段）在 /special 响应中可用且正确。"""
    t = _today()
    _mk_project(db_session, "专项延期", status="进行中", is_special=True, plan_end=t - timedelta(days=5))
    _mk_project(db_session, "专项到期", status="进行中", is_special=True, plan_end=t + timedelta(days=3))
    _mk_project(db_session, "专项无阶段", status="进行中", is_special=True, plan_end=t + timedelta(days=30))
    db_session.commit()

    resp = client.get("/api/projects/special")
    assert resp.status_code == 200
    by_name = {p["name"]: p for p in resp.json()}

    # 延期：plan_end < today 且未完成（前端 🔴 触发条件的数据就绪）
    assert by_name["专项延期"]["plan_end"] == (t - timedelta(days=5)).isoformat()
    assert by_name["专项延期"]["status"] == "进行中"
    # 即将到期：plan_end 在 7 天内（前端 🟡 触发条件的数据就绪）
    assert by_name["专项到期"]["plan_end"] == (t + timedelta(days=3)).isoformat()
    # 无阶段：phases 为空（前端 ⚠️ 触发条件的数据就绪）
    assert by_name["专项无阶段"]["phases"] == []


def test_7_is_special_toggle_effective(client, db_session):
    """用例 7：创建/更新 is_special 开关生效、列表正确过滤。"""
    # 创建 is_special=true
    resp = client.post("/api/projects", json={
        "name": "开关项目", "category": "新需求", "owner": "张三", "market": "拉美区",
        "is_special": True,
    })
    assert resp.status_code == 201, resp.text
    pid = resp.json()["id"]
    assert resp.json()["is_special"] is True

    # 专项列表有、普通列表无
    assert [p["id"] for p in client.get("/api/projects/special").json()] == [pid]
    assert all(p["id"] != pid for p in client.get("/api/projects").json())

    # 更新取消开关 → 进普通列表、出专项列表
    resp = client.put(f"/api/projects/{pid}", json={"is_special": False})
    assert resp.status_code == 200, resp.text
    assert resp.json()["is_special"] is False
    assert all(p["id"] != pid for p in client.get("/api/projects/special").json())
    assert any(p["id"] == pid for p in client.get("/api/projects").json())

    # 普通项目更新为 is_special=true → 反向迁移
    resp = client.put(f"/api/projects/{pid}", json={"is_special": True})
    assert resp.status_code == 200
    assert [p["id"] for p in client.get("/api/projects/special").json()] == [pid]
    assert all(p["id"] != pid for p in client.get("/api/projects").json())


def test_8_normal_list_and_dashboard_isolated(client, db_session):
    """用例 8：普通列表与 Dashboard 隔离（is_special=true 不出现在普通列表与 Dashboard 统计）。"""
    t = _today()
    # 普通：延期项目 + 进行中阶段（有返工）
    p_norm = _mk_project(db_session, "普通统计项目", status="进行中", plan_end=t - timedelta(days=2))
    ph_norm = _mk_phase(db_session, p_norm, "普通延期阶段", t - timedelta(days=10), t - timedelta(days=1))
    ph_norm.rework_count = 2
    # 专项：同样延期 + 阶段有返工
    sp = _mk_project(db_session, "专项统计项目", status="进行中", is_special=True, plan_end=t - timedelta(days=2))
    ph_sp = _mk_phase(db_session, sp, "专项延期阶段", t - timedelta(days=10), t - timedelta(days=1))
    ph_sp.rework_count = 3
    db_session.commit()

    # 普通列表
    names = [p["name"] for p in client.get("/api/projects").json()]
    assert "普通统计项目" in names and "专项统计项目" not in names

    # Dashboard 统计
    stats = client.get("/api/dashboard/stats").json()
    assert stats["total_projects"] == 1
    assert [d["name"] for d in stats["delayed_projects"]] == ["普通统计项目"]
    assert {s["status"] for s in stats["project_status"]} == {"进行中"}
    assert stats["total_rework_count"] == 2
    assert [r["project_name"] for r in stats["rework_phases"]] == ["普通统计项目"]
    assert [d["project_name"] for d in stats["delayed_phases"]] == ["普通统计项目"]


def test_9_detail_phase_permissions(client, db_session):
    """用例 9：非 admin/manager 访问专项详情/甘特/阶段 → 403；普通项目不受影响。

    engineer 已关联 resource 并分配到专项/普通阶段——若没有专项守卫，check_phase_access
    会放行工程师改被分配的阶段（200）；守卫存在时专项阶段一律 403，普通阶段维持 200。
    """
    t = _today()
    sp = _mk_project(db_session, "权限专项", is_special=True)
    sp_ph = _mk_phase(db_session, sp, "专项权限阶段", t, t + timedelta(days=5))
    np = _mk_project(db_session, "权限普通")
    np_ph = _mk_phase(db_session, np, "普通权限阶段", t, t + timedelta(days=5))
    r_perm = _mk_resource(db_session, "权限人")
    sp_ph.assignees = [r_perm]
    np_ph.assignees = [r_perm]
    _mk_user(db_session, "eng_perm", "engineer", "工程师乙")
    db_session.commit()
    db_session.query(User).filter_by(username="eng_perm").first().resource_id = r_perm.id
    db_session.commit()

    _login(client, "eng_perm")

    # 专项：详情/甘特/阶段列表/阶段详情/阶段 CRUD 全 403
    assert client.get(f"/api/projects/{sp.id}").status_code == 403
    assert client.get(f"/api/projects/{sp.id}/gantt").status_code == 403
    assert client.get(f"/api/projects/{sp.id}/critical-path").status_code == 403
    assert client.get(f"/api/projects/{sp.id}/phases").status_code == 403
    assert client.get(f"/api/phases/{sp_ph.id}").status_code == 403
    assert client.put(f"/api/phases/{sp_ph.id}", json={"status": "已完成"}).status_code == 403
    assert client.post(f"/api/phases/{sp_ph.id}/move", params={"direction": "up"}).status_code == 403
    assert client.delete(f"/api/phases/{sp_ph.id}").status_code == 403
    assert client.post(f"/api/phases/{sp_ph.id}/rework",
                       json={"to_status": "未开始", "reason": "测试"}).status_code == 403
    assert client.post(f"/api/projects/{sp.id}/phases", json={
        "phase_type": "P5", "name": "新阶段", "sequence": 2,
    }).status_code == 403

    # 普通项目不受影响：详情/甘特/阶段列表 200；被分配工程师改普通阶段 200
    assert client.get(f"/api/projects/{np.id}").status_code == 200
    assert client.get(f"/api/projects/{np.id}/gantt").status_code == 200
    assert client.get(f"/api/projects/{np.id}/phases").status_code == 200
    resp = client.put(f"/api/phases/{np_ph.id}", json={"progress": 60})
    assert resp.status_code == 200, resp.text

    # viewer 同样 403
    _mk_user(db_session, "view_perm", "viewer", "观察员乙")
    db_session.commit()
    _login(client, "view_perm")
    assert client.get(f"/api/projects/{sp.id}").status_code == 403
    assert client.get(f"/api/projects/{sp.id}/gantt").status_code == 403
    assert client.get(f"/api/projects/{np.id}").status_code == 200


def test_10_list_aggregations_excluded(client, db_session):
    """用例 10：列表聚合排除：搜索/看板（GET /api/projects）无专项；独立聚合接口（Excel 导出）核对。"""
    _mk_project(db_session, "聚合专项", is_special=True)
    _mk_project(db_session, "聚合普通")
    db_session.commit()

    # 看板/搜索走 GET /api/projects：无专项（带筛选也一致）
    all_names = [p["name"] for p in client.get("/api/projects").json()]
    assert "聚合专项" not in all_names and "聚合普通" in all_names
    filtered = [p["name"] for p in client.get("/api/projects", params={"category": "新需求"}).json()]
    assert "聚合专项" not in filtered

    # 独立聚合接口：Excel 导出不含专项项目
    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
    ws = wb.active
    exported = " ".join(str(row[2].value) for row in ws.iter_rows(min_row=3) if len(row) > 2 and row[2].value)
    assert "聚合普通" in exported
    assert "聚合专项" not in exported


def test_11_project_level_permissions(client, db_session):
    """用例 11：非 admin/manager 更新/删除专项项目、创建 is_special=true → 403。"""
    sp = _mk_project(db_session, "写权限专项", is_special=True)
    db_session.commit()

    _mk_user(db_session, "eng_write", "engineer", "工程师丙")
    _mk_user(db_session, "mgr_write", "manager", "经理乙")
    db_session.commit()

    # engineer：更新/删除专项项目 403；创建 is_special=true 403（创建本就需要 admin/manager）
    _login(client, "eng_write")
    assert client.put(f"/api/projects/{sp.id}", json={"name": "改名"}).status_code == 403
    assert client.delete(f"/api/projects/{sp.id}").status_code == 403
    assert client.post("/api/projects", json={
        "name": "工程师建专项", "category": "新需求", "owner": "张三", "market": "拉美区",
        "is_special": True,
    }).status_code == 403

    # manager：更新专项项目 200（专项项目不限于本人负责）；删除仍走申请流程 403
    _login(client, "mgr_write")
    resp = client.put(f"/api/projects/{sp.id}", json={"name": "经理改专项名"})
    assert resp.status_code == 200, resp.text
    assert client.delete(f"/api/projects/{sp.id}").status_code == 403
    # manager 创建 is_special=true → 201
    resp = client.post("/api/projects", json={
        "name": "经理建专项", "category": "新需求", "owner": "经理乙", "market": "拉美区",
        "is_special": True,
    })
    assert resp.status_code == 201, resp.text
    assert resp.json()["is_special"] is True


def test_12_badge_boundaries_backend(client, db_session):
    """用例 12：角标边界（后端可测部分）：搁置/已完成专项无角标数据；plan_end 为空不触发延期/到期。"""
    t = _today()
    _mk_project(db_session, "搁置专项", status="搁置", is_special=True, plan_end=t - timedelta(days=10))
    _mk_project(db_session, "完成专项", status="已完成", is_special=True, plan_end=t - timedelta(days=3))
    _mk_project(db_session, "无期限专项", status="进行中", is_special=True, plan_end=None)
    db_session.commit()

    resp = client.get("/api/projects/special")
    assert resp.status_code == 200
    by_name = {p["name"]: p for p in resp.json()}

    # 搁置/已完成：数据上 plan_end 已过期但不触发角标（前端按 status 抑制）
    assert by_name["搁置专项"]["status"] == "搁置"
    assert by_name["完成专项"]["status"] == "已完成"
    # plan_end 为空：不触发延期/到期（字段为 null）
    assert by_name["无期限专项"]["plan_end"] is None
    assert by_name["无期限专项"]["status"] == "进行中"


# ---- §七 用例 13/14：常规导入隔离 + 专项导入（SPECIAL_PROJECT §五·B，评审 🔴 处置）----


def _make_workbook(rows: list[list]) -> bytes:
    """按 14 列新格式构造 Excel（与 test_merge_import 同构）。每行: [编号, 类目, 名称,
    负责人, 市场, 阶段类型, 计划开始, 计划结束, 实际开始, 实际结束,
    阶段负责人, 阶段状态, 阶段进度, 备注]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目填报"
    headers = ["项目编号", "项目类目", "项目名称", "项目负责人", "市场", "阶段类型",
               "计划开始", "计划结束", "实际开始", "实际结束",
               "阶段负责人", "阶段状态", "阶段进度", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            ws.cell(3 + i, c, v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_old_workbook(rows: list[list]) -> bytes:
    """按 8 列旧格式构造 Excel。每行: [编号, 类目, 名称, 负责人, 计划开始, 计划结束, 状态, 交接人]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目情况统计-海外"
    headers = ["项目编号", "项目类目", "项目名称", "负责人", "计划开始", "计划结束", "状态", "交接人"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(3 + i, c, v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mk_special_with_phase(db_session, name: str) -> tuple[Project, Phase, Resource]:
    """专项项目 + 1 阶段（assignee 关联到独立 Resource），返回 (project, phase, resource)。"""
    sp = _mk_project(db_session, name, is_special=True)
    r = _mk_resource(db_session, f"{name}-执行人")
    ph = _mk_phase(db_session, sp, f"{name}-阶段", _today(), _today() + timedelta(days=30))
    ph.assignees = [r]
    return sp, ph, r


def test_13_regular_replace_preserves_special(client, db_session):
    """用例 13a：常规全量导入（replace）后专项项目（阶段/assignee/Resource）原样保留；报告含专项提示。"""
    sp, sp_ph, sp_res = _mk_special_with_phase(db_session, "专项保留项目")
    # 常规项目 + 仅被常规引用的 Resource（应被清掉）
    np = _mk_project(db_session, "常规旧项目")
    np_res = _mk_resource(db_session, "常规旧人员")
    _mk_phase(db_session, np, "常规旧阶段", _today(), _today() + timedelta(days=10)).assignees = [np_res]
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "常规新项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    resp = client.post("/api/import/excel", files={"file": ("t.xlsx", data)},
                       params={"mode": "replace"})
    assert resp.status_code == 200, resp.text
    report = resp.json()

    # 报告注明专项不受影响
    assert any("专项项目 1 个不受影响" in w["message"] for w in report["warnings"])

    # 专项项目原样保留（id/阶段/assignee Resource 均未动）
    db_session.expire_all()
    assert db_session.query(Project).filter(Project.is_special.is_(True)).count() == 1
    assert db_session.query(Phase).filter(Phase.id == sp_ph.id).count() == 1
    kept_ph = db_session.query(Phase).filter(Phase.id == sp_ph.id).one()
    assert [r.name for r in kept_ph.assignees] == ["专项保留项目-执行人"]
    assert db_session.query(Resource).filter(Resource.id == sp_res.id).count() == 1

    # 常规域被清空重建：旧常规项目/阶段/人员没了，新项目在（专项仍保留）
    assert db_session.query(Project).filter(Project.name == "常规旧项目").count() == 0
    assert db_session.query(Resource).filter(Resource.name == "常规旧人员").count() == 0
    names = [p.name for p in db_session.query(Project).filter(Project.is_special.is_(False)).all()]
    assert names == ["常规新项目"]


def test_13_regular_merge_not_into_special(client, db_session):
    """用例 13b：常规合并导入与专项同名 → 不合并进专项（新建常规项目），专项原样。"""
    sp, sp_ph, _ = _mk_special_with_phase(db_session, "同名专项")
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "同名专项", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "工业设计", "2026-07-01", "2026-07-10", "", "", "李四", "进行中", 50, ""],
    ])
    resp = client.post("/api/import/excel", files={"file": ("t.xlsx", data)},
                       params={"mode": "merge"})
    assert resp.status_code == 200, resp.text
    report = resp.json()

    # 不合并进专项 → 视为新建常规项目
    assert report["projects_created"] == 1
    assert report["projects_updated"] == 0

    # 专项原样（阶段未被合并/新增）
    db_session.expire_all()
    assert db_session.query(Project).filter(Project.name == "同名专项", Project.is_special.is_(True)).count() == 1
    assert db_session.query(Phase).filter(Phase.id == sp_ph.id).count() == 1

    # 新建的常规项目（同名）存在
    reg = db_session.query(Project).filter(Project.is_special.is_(False), Project.name == "同名专项").first()
    assert reg is not None
    assert len(reg.phases) == 1


def test_13_regular_preview_excludes_special(client, db_session):
    """用例 13c：常规预览专项不计入 existing/matched/kept（专项域数据全不出现）。"""
    sp, sp_ph, sp_res = _mk_special_with_phase(db_session, "专项预览项目")
    np = _mk_project(db_session, "常规预览项目")
    np_res = _mk_resource(db_session, "常规预览人员")
    _mk_phase(db_session, np, "常规预览阶段", _today(), _today() + timedelta(days=10)).assignees = [np_res]
    db_session.commit()

    # 文件只含一个全新项目（防传错对比：matched=0/new=1/missing=1，全部常规域口径）
    data = _make_workbook([
        [1, "新需求", "常规新项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    resp = client.post("/api/import/preview", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text
    pv = resp.json()

    # existing：仅常规域（专项项目/阶段/专项引用 Resource 不计入"将被清空"）
    assert pv["existing"]["projects"] == 1
    assert pv["existing"]["phases"] == 1
    assert pv["existing"]["resources"] == 1  # 仅常规预览人员；专项执行人保留

    # match/kept 均不含专项
    assert pv["match"] == {"matched": 0, "new": 1, "missing": 1}
    assert pv["kept_count"] == 1
    assert all(p["name"] != "专项预览项目" for p in pv["projects_preview"])


def test_13_special_numeric_code_does_not_block_replace(client, db_session):
    """编号冲突回归（实测 PG 复现）：专项项目占用纯数字编号 '1' 时，
    常规全量导入续编不撞唯一约束 project.code，专项保留。"""
    sp = _mk_project(db_session, "编号专项", is_special=True)
    sp.code = "1"  # 专项导入的文件行号编号方式（'1'..'N'）
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "常规新项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    resp = client.post("/api/import/excel", files={"file": ("t.xlsx", data)},
                       params={"mode": "replace"})
    assert resp.status_code == 200, resp.text

    db_session.expire_all()
    sp2 = db_session.query(Project).filter(Project.name == "编号专项").one()
    assert sp2.code == "1"
    reg = db_session.query(Project).filter(Project.name == "常规新项目").one()
    assert reg.code == "2"  # 续编：现有最大编号 + 1


def test_14_special_import_resets_special_domain(client, db_session):
    """用例 14a：专项导入全量重置专项域——旧专项删、新专项按文件建 is_special=True、
    阶段类型自由文本直存、FS 链建好、常规项目不受影响；最近报告槽位更新。"""
    old_sp, old_ph, _ = _mk_special_with_phase(db_session, "旧专项")
    np = _mk_project(db_session, "常规不动项目")
    np_ph = _mk_phase(db_session, np, "常规阶段", _today(), _today() + timedelta(days=10))
    db_session.commit()

    data = _make_workbook([
        [1, "专项类目", "新专项A", "专项负责人", "拉美区", "", "2026-07-01", "2026-09-30", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "电磁兼容测试", "2026-07-01", "2026-07-20", "", "", "专项执行人", "进行中", 30, "专项备注"],
        ["1-2", "", "", "", "", "可靠性验证", "2026-07-21", "2026-08-10", "", "", "专项执行人", "未开始", 0, ""],
    ])
    resp = client.post("/api/import/special", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["projects_imported"] == 1
    assert report["phases_imported"] == 2
    assert report["errors"] == []

    # 旧专项删除（含阶段级联），新专项 is_special=True、阶段类型原样存储
    db_session.expire_all()
    assert db_session.query(Project).filter(Project.name == "旧专项").count() == 0
    assert db_session.query(Phase).filter(Phase.name == "旧专项-阶段").count() == 0
    sp = db_session.query(Project).filter(Project.name == "新专项A").first()
    assert sp is not None and sp.is_special is True
    assert sp.owner == "专项负责人" and sp.market == "拉美区" and sp.category == "专项类目"
    ph1 = next(p for p in sp.phases if p.sequence == 1)
    ph2 = next(p for p in sp.phases if p.sequence == 2)
    assert ph1.phase_type == "电磁兼容测试"
    assert ph2.phase_type == "可靠性验证"
    assert ph1.remark == "专项备注"
    # assignee 全局复用 Resource（同名人员不重复建）
    res = db_session.query(Resource).filter(Resource.name == "专项执行人").one()
    assert res.id == ph1.assignees[0].id == ph2.assignees[0].id
    # FS 串联链建好（2 阶段 1 条链）
    from app.models import Dependency
    deps = db_session.query(Dependency).filter(
        Dependency.from_phase_id == ph1.id, Dependency.to_phase_id == ph2.id
    ).count()
    assert deps == 1

    # 常规项目完全不受影响（id/阶段原样）
    np2 = db_session.query(Project).filter(Project.name == "常规不动项目").one()
    assert np2.is_special is False
    assert db_session.query(Phase).filter(Phase.id == np_ph.id).count() == 1

    # 最近报告槽位 = 专项报告
    last = client.get("/api/import/report").json()
    assert last["projects_imported"] == 1 and last["phases_imported"] == 2


def test_14_special_import_renumbers_after_regular_codes(client, db_session):
    """编号冲突回归：常规项目占用纯数字编号 '1' 时，专项导入续编不撞唯一约束。"""
    np = _mk_project(db_session, "常规占号")
    np.code = "1"
    db_session.commit()

    data = _make_workbook([
        [1, "新需求", "专项新项目", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])
    resp = client.post("/api/import/special", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text

    db_session.expire_all()
    sp = db_session.query(Project).filter(Project.name == "专项新项目").one()
    assert sp.is_special is True and sp.code == "2"  # 续编：现有最大编号 + 1


def test_14_special_import_permissions(client, db_session):
    """用例 14b：专项导入/预览仅 admin/manager——engineer/viewer 403；manager 200。"""
    data = _make_workbook([
        [1, "新需求", "权限专项导入", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
    ])

    # engineer / viewer → 403（预览 + 导入都要拦）
    _mk_user(db_session, "eng_simp", "engineer", "工程师丁")
    _mk_user(db_session, "view_simp", "viewer", "观察员丙")
    db_session.commit()
    _login(client, "eng_simp")
    assert client.post("/api/import/special-preview",
                       files={"file": ("t.xlsx", data)}).status_code == 403
    assert client.post("/api/import/special",
                       files={"file": ("t.xlsx", data)}).status_code == 403
    _login(client, "view_simp")
    assert client.post("/api/import/special-preview",
                       files={"file": ("t.xlsx", data)}).status_code == 403
    assert client.post("/api/import/special",
                       files={"file": ("t.xlsx", data)}).status_code == 403

    # manager → 200（预览 + 导入）
    _mk_user(db_session, "mgr_simp", "manager", "经理丙")
    db_session.commit()
    _login(client, "mgr_simp")
    resp = client.post("/api/import/special-preview", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text
    assert resp.json()["incoming"]["projects"] == 1
    resp = client.post("/api/import/special", files={"file": ("t.xlsx", data)})
    assert resp.status_code == 200, resp.text
    assert resp.json()["projects_imported"] == 1
    assert db_session.query(Project).filter(Project.is_special.is_(True)).count() == 1


def test_14_parse_special_mode_raw_phase_type(db_session):
    """用例 14c（单元）：parse_workbook(special=True) 阶段类型列原样存储、无警告；
    旧格式无类型列 → phase_type 留空；对照：常规模式同样内容会警告+跳过。"""
    from app.services.excel_importer import parse_workbook

    # 新格式：任意自由文本阶段类型
    data = _make_workbook([
        [1, "新需求", "专项解析A", "张三", "拉美区", "", "2026-07-01", "2026-08-01", "", "", "", "", "", ""],
        ["1-1", "", "", "", "", "电磁兼容测试", "2026-07-01", "2026-07-20", "", "", "李四", "进行中", 30, ""],
    ])
    parsed = parse_workbook(data, special=True)
    assert len(parsed.projects) == 1
    assert parsed.projects[0].phases[0].phase_type == "电磁兼容测试"
    assert parsed.projects[0].phases[0].name == "电磁兼容测试"
    assert parsed.report.errors == [] and parsed.report.warnings == []

    # 对照：常规模式同一文件 → 任意类型不可解析，警告 + 跳过该行
    parsed_reg = parse_workbook(data)
    assert parsed_reg.projects[0].phases == []
    assert any("无法识别" in w.message for w in parsed_reg.report.warnings)

    # 旧格式：无阶段类型列 → phase_type 留空（不映射、不 error）；阶段名走名称列
    old = _make_old_workbook([
        [1, "新需求", "旧格式专项", "张三", "", "", "进行中", ""],
        ["1-1", "", "电磁兼容测试", "李四", "", "", "已完成", ""],
    ])
    parsed_old = parse_workbook(old, special=True)
    assert len(parsed_old.projects) == 1
    ph = parsed_old.projects[0].phases[0]
    assert ph.phase_type == "" and ph.name == "电磁兼容测试"
    assert ph.assignees == ["李四"]
    assert parsed_old.report.errors == []
